#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'Curry'
'''
description:
    一、项目名称：U3项目自动化测试
    二、涵盖测试模块：
        1、云卡和实体卡切换
        2、Flash
        3、显示
        4、加入云卡后的稳定性
'''

import pytest, subprocess, time, os, allure, sys, random
# 获取项目路径
BASEDIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
sys.path.append(BASEDIR)
from common import models, logger
from common.ui_operation import S20i_uiOpration
from config import settings, U3_settings
import uiautomator2 as u2
# from pywinauto import application
from openpyxl import load_workbook

# 获取项目名称
project_name = os.path.basename(__file__).split('.')[0]
# 创建log对象
log_name = '.'.join(['_'.join([project_name, time.strftime('%Y%m%d%H%M%S', time.localtime())]), "log"])
log_file = os.path.join(BASEDIR, "logs", log_name)
log = logger.Logger(screen_output=True, log_level='logging.INFO', log_file=log_file).create_logger()

def screenShot(d, title):
    '''
    在uiautomator2的截图函数上封装的截图方法
    :param
        d:uiautomator2对象
        title: 自定义截图的名称
    :return: 返回截图的路径
    '''
    screenShot_file = '/'.join([test_record_path, '.'.join([title, 'png'])])
    d.screenshot(screenShot_file)
    return screenShot_file

def pull_ScreenRecord(d, title):
    list = d.adb_shell('ls /sdcard/ScreenRecord/record/')
    ScreenRecord_list = list.strip().split('\n')
    file = '/'.join(['/sdcard/ScreenRecord/record', ScreenRecord_list[-1]])
    save_file = '/'.join([test_record_path, '.'.join([title, 'mp4'])])
    d.pull(src=file, dst=save_file)
    return save_file

@allure.feature('云卡和实体卡切换')
class Test_CloudSimAndPhysicalSimSwitch:
    # 测试类起始函数
    def setup_class(self):
        # self.start_SwitchComTool = application.Application().start(settings.Qualcomm_SwitchCom_file)    # 启动高通自动切口工具
        log.info("-"*20 + "Start Test_CloudSimAndPhysicalSimSwitch" + "-"*20)
        self.u3 = models.U3(device_id=U3_settings.test_device_info['id'], log_project=log)
        self.s20i = models.Glocalme(device_id=U3_settings.auxiliary_device_info['id'], log_project=log)
        self.u3.wait_device_connect()    # 等待测试设备连接
        self.s20i.wait_device_connect()    # 等待辅助设备连接
        # 执行uiautomator2的init操作，在测试设备端安装uiautomator app，minicap和minitouch，atx-agent
        uiautomator2_init = subprocess.getoutput("python -m uiautomator2 init")
        # self.u3_d = u2.connect(U3_settings.test_device_info['id'])
        self.s20i_d = u2.connect('c58790e3')
        self.s20i.wakey()    # 设置辅助机屏幕常亮
        # 判断辅助机屏幕是否锁屏状态，如果锁屏就给设备解锁（仅适用滑动解锁）
        if self.s20i_d(resourceId="com.android.systemui:id/lock_icon").exists():
            self.s20i_d.xpath('//android.widget.FrameLayout').click()
            self.s20i_d(resourceId="com.android.systemui:id/lock_icon").click()
        version = self.u3.get_current_version()    # 获取测试设备的版本
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)  # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path, '.'.join([version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', '整机测试内容_无屏MIFI.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["云卡和实体卡切换"]  # 选择表

    def setup(self):
        for i in range(2):
            self.s20i_d.press('home')
        self.s20i_d(resourceId="com.android.systemui:id/recent_apps").click()
        if self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").exists(3):
            self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").click()
            log.info("cleared background app success.")
        else:
            log.info('no background app running.')
        self.s20i_d.press('home')

    # @pytest.mark.skip(msg="test")
    def test01(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step('第二步：测试设备关机，插入实体卡后开机，登陆webui界面查看是否显示选择登陆界面'):
            while True:
                inp = input('\033[35;0m%s' % '请将测试设备关机，并插入实体卡后开机，确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.wait_device_connect()
            self.open_webui()
            test_result_01 = pytest.assume(self.s20i_d(text="A new physical SIM is detected, please select the way you want to use mobile data.").exists(10))
            if not test_result_01:
                self.ws1['G2'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test01_Step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test01 step2 failed screenshot', allure.attachment_type.PNG)
                assert False
        with allure.step('第三步：选择实体卡登陆'):
            self.s20i_d(resourceId="text_sim_1").click()    # 选择sim卡
            self.s20i_d(text="Apply").click()
            test_result_02 = pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use SIM 1(Manual) for Internet.").exists(10))    # 判断是否弹窗提示使用实体卡
            self.s20i_d(resourceId="notice_btn_ok").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_03 = pytest.assume(self.u3.login_type() == 0)    # 判断测试机是否使用实体卡
            if all([test_result_02, test_result_03]):
                self.ws1['G2'] = "PASS"
            else:
                self.ws1['G2'] = "FAIL"
                log.error('测试结果test_result_01:%stest_result_02：%s' % (test_result_01, test_result_02))

    def test02(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step('第二步：测试设备关机，插入实体卡后开机'):
            while True:
                inp = input('\033[35;0m%s' % '请将测试设备关机, 并插入实体卡后开机，确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.wait_device_connect()
            time.sleep(3)
        with allure.step('第三步：辅助机连接测试机wifi，打开浏览器进入webUi界面，选择云卡登陆'):
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()    # 选择云卡
            self.s20i_d(text="Apply").click()
            test_result_01 = pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use Cloud SIM(Manual) for Internet.").exists(10))    # 判断是否弹窗提示使用云卡
            self.s20i_d(resourceId="notice_btn_ok").click()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(5)
            test_result_02 = pytest.assume(self.u3.login_type() == 1)    # 判断测试机是否使用云卡
            if not all([test_result_01, test_result_02]):
                self.ws1['G3'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='Step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, '开机云卡登陆失败webui的截图', allure.attachment_type.PNG)
                assert False
        with allure.step("第四步: 切换SIM卡登陆"):
            self.login_webui()
            # 进入SIM卡管理界面，切换到SIM卡登陆
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_1").click()
            self.s20i_d(text="Apply").click()
            test_result_01 = pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use SIM 1(Manual) for Internet.").exists(10))  # 判断是否弹窗提示使用实体卡
            self.s20i_d(resourceId="notice_btn_ok").click()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(5)
            test_result_02 = pytest.assume(self.u3.login_type() == 0)  # 判断测试机是否使用实体卡
            if not all([test_result_01, test_result_02]):
                self.ws1['G3'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='Step3_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, '切换云卡登陆失败webui的截图', allure.attachment_type.PNG)
                assert False
        with allure.step("第五步：切换到云卡登陆"):
            time.sleep(90)    # 此处设置90秒等待，以免切换卡太频繁导致无法切换
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()
            self.s20i_d(text="Apply").click()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(5)
            test_result_03 = pytest.assume(self.u3.login_type() == 1)  # 判断测试机是否使用云卡
            if not test_result_03:
                self.ws1['G3'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='Step4_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, '切换云卡登陆失败webui的截图', allure.attachment_type.PNG)
            else:
                self.ws1['G3'] = "PASS"

    # @pytest.mark.skip()
    def test03(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step('第一步：测试设备关机，拨出实体卡后开机'):
            inp = input('\033[35;0m请将测试设备关机，拔出实体卡后开机，确认操作完成后输入y继续：')
            if inp == 'y':
                self.u3.wait_device_connect()
                self.u3.wait_network_connect()
                time.sleep(1)
            else:
                log.error('用户操作失败，结束测试, inp:%s' % inp)
                self.ws1['G4'] = "NONE"
                assert False
        with allure.step("第二步：插入SIM卡"):
            inp = input('\033[35;0m请插入实体卡，确认操作完成后输入y继续：')
            if inp == 'y':
                self.u3.wait_device_connect()
            else:
                log.error('用户操作失败，结束测试, inp:%s' % inp)
                self.ws1['G4'] = "NONE"
                assert False
        with allure.step("第三步：辅助机连接测试机wifi，登陆webui界面切换SIM卡"):
            time.sleep(30)
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            # 点击选择SIM 1，查看是否可以设置为SIM 1，此动作重复两次
            for i in range(2):
                self.s20i_d(resourceId="text_sim_1").click()
                time.sleep(1)
            sim_selected = self.s20i_d.xpath('//*[@resource-id="span_radio_state_1"]/android.view.View[2]').exists    # 判断是否可以选择实体卡登陆，return True or False
            # global test03_result
            test03_result = pytest.assume(sim_selected == False)
            if test03_result:
                self.ws1['G4'] = "PASS"
            else:
                self.ws1['G4'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='Step4_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, '实体卡切换按钮置灰失败的截图', allure.attachment_type.PNG)

    def test04(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入SIM卡，重启，选择云卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(3)
            self.u3.wait_device_connect()
            time.sleep(30)
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_01 = pytest.assume(self.u3.login_type() == 1)  # 判断测试机是否使用云卡
            if test_result_01:
                self.ws1['G5'] = "PASS"
            else:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test04_step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test04 cloudsim login failed ScreenShot', allure.attachment_type.PNG)
                self.ws1['G5'] = "FAIL"


    def test05(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入SIM卡，重启，选择云卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(3)
            self.u3.wait_device_connect()
            time.sleep(30)
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_01 = pytest.assume(self.u3.login_type() == 1)  # 判断测试机是否使用云卡
            if not test_result_01:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='Step5_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test05 change cloudSim failed ScreenShot', allure.attachment_type.PNG)
                self.ws1['G6'] = "FAIL"
                assert False
        with allure.step("第三步：拔插sim卡，登陆webui界面切换SIM卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插拔实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_1").click()  # 选择实体卡
            self.s20i_d(text="Apply").click()
            test_result_02 = pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use SIM 1(Manual) for Internet.").exists(10))  # 判断是否弹窗提示使用实体卡
            if not test_result_02:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test05_Step3_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test05 step3 failed ScreenShot', allure.attachment_type.PNG)
            self.s20i_d(resourceId="notice_btn_ok").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_03 = pytest.assume(self.u3.login_type() == 0)  # 判断测试机是否使用实体卡
            if all([test_result_02, test_result_03]):
                self.ws1['G6'] = "PASS"
            else:
                self.ws1['G6'] = "FAIL"

    def test06(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：选择云卡登陆，云卡登陆过程中拔插SIM卡"):
            if self.u3.login_type() == 0:
                self.open_webui()
                self.login_webui()
                self.s20i_d(text="Settings").click()
                self.s20i_d(text="SIM card management").click()
                self.s20i_d(resourceId="sim_select").click()
                self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
                self.s20i_d(text="Apply").click()
                time.sleep(3)
            elif self.u3.login_type() == 1:
                self.u3.reboot()
                self.u3.wait_device_disconnect()
                self.u3.wait_device_connect()
            # pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use Cloud SIM(Manual) for Internet.").exists(10))  # 判断是否弹窗提示使用云卡
            # self.s20i_d(resourceId="notice_btn_ok").click()
            # self.u3.wait_network_connect()
            while not self.u3.connect_network():
                inp = input('\033[35;0m%s' % "请在云卡登陆成功之前拔插sim卡，确认操作完成后输入y继续：")
                if inp == 'y':
                    if not self.u3.connect_network():
                        with allure.step("第二步：云卡登陆过程中，进入webui切换SIM卡登陆"):
                            for i in range(2):
                                self.s20i_d.press("home")
                            self.open_webui()
                            self.login_webui()
                            self.s20i_d(text="Settings").click()
                            self.s20i_d(text="SIM card management").click()
                            self.s20i_d(resourceId="sim_select").click()
                            if not self.u3.connect_network():
                                self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
                                self.s20i_d(text="Apply").click()
                                break
                            else:
                                log.error("未在云卡登陆成功之前切换SIM卡，重启设备后重新执行一次。")
                                self.u3.reboot()
                                time.sleep(3)
                                self.u3.wait_device_connect()
                                time.sleep(1)
                                continue
                    else:
                        log.error("未在云卡登陆成功之前拔插SIM卡，重启设备后重新执行一次。")
                        self.u3.reboot()
                        time.sleep(3)
                        self.u3.wait_device_connect()
                        continue
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            test_result_01 = pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use SIM 1(Manual) for Internet.").exists(10))  # 判断是否弹窗提示使用实体卡
            if not test_result_01:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='Step6_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test06 change physical Sim failed ScreenShot', allure.attachment_type.PNG)
            self.s20i_d(resourceId="notice_btn_ok").click()
            self.u3.wait_network_connect()
            time.sleep(5)
            test_result_02 = pytest.assume(self.u3.login_type() == 0)  # 判断测试机是否使用实体卡
            if all([test_result_01, test_result_02]):
                self.ws1['G7'] = "PASS"
            else:
                self.ws1['G7'] = "FAIL"

    def test07(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入SIM卡，重启，选择云卡登陆"):
            while True:
                inp01 = input('\033[35;0m请插入实体卡，确认操作完成后输入y继续(输入q退出测试)：')
                if inp01 == 'y':
                    break
                elif inp01 == 'q':
                    log.error('用户退出测试，结束测试, inp01:%s' % inp01)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp01)
                    continue
            self.u3.reboot()
            time.sleep(3)
            self.u3.wait_device_connect()
            time.sleep(30)
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            test_result_01 = pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use Cloud SIM(Manual) for Internet.").exists(10))  # 判断是否弹窗提示使用云卡
            self.s20i_d(resourceId="notice_btn_ok").click()
            self.u3.wait_network_connect()
            time.sleep(5)
            test_result_02 = pytest.assume(self.u3.login_type() == 1)  # 判断测试机是否使用云卡
            if not all([test_result_01, test_result_02]):
                self.ws1['G8'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='Step8_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test07 change cloudSim failed ScreenShot', allure.attachment_type.PNG)
                assert False
        with allure.step("第三步：拔插SIM卡，重启设备"):
            while True:
                inp02 = input('\033[35;0m请拔插实体卡，确认操作完成后输入y继续(输入q退出测试)：')
                if inp02 == 'y':
                    break
                elif inp02 == 'q':
                    log.error('用户退出测试，结束测试, inp02:%s' % inp02)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp02)
                    continue
            self.u3.reboot()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(5)
            test_result = pytest.assume(self.u3.login_type() == 1)  # 判断测试机是否使用云卡
            if test_result:
                self.ws1['G7'] = "PASS"
            else:
                self.ws1['G7'] = "FAIL"

    def test08(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：判断当前是否使用云卡，如果不是云卡切换到云卡"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入带pin码的SIM卡，重启设备，登陆webui界面"):
            while True:
                inp = input('\033[35;0m%s' % '请插入带pin码的SIM卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.open_webui()
            test_result_01 = pytest.assume(self.s20i_d(resourceId="text_sim_1", text="SIM 1(PIN required)").exists(10))    # 判断是否出现选择登陆界面
            if not test_result_01:
                self.ws1['G9'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test08_step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test08 step2 failed ScreenShot', allure.attachment_type.PNG)
                assert False
        with allure.step("第三步：切换SIM卡登陆"):
            self.s20i_d(resourceId="text_sim_1", text="SIM 1(PIN required)").click()
            self.s20i_d(text="Apply").click()
            test_result_02 = pytest.assume(self.s20i_d(resourceId="notice_cnt", text="You will use SIM 1(PIN required) for Internet.").exists(10))    # 判断是否弹窗提示使用SIM 1
            time.sleep(3)
            test_result_03 = pytest.assume(self.u3.connect_network() == False)
            if all([test_result_02, test_result_03]):
                self.ws1['G9'] = "PASS"
            else:
                self.ws1['G9'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test08_step3_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test08 step3 change SIM 1 failed ScreenShot', allure.attachment_type.PNG)

    def test09(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入实体卡后重启设备，选择云卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()
            self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_01 = pytest.assume(self.u3.login_type() == 1)
        with allure.step("第三步：拔掉SIM卡，登陆webui界面切换SIM卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
            self.s20i_d(text="Apply").click()
            time.sleep(30)
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()    # 点击搜索框
            self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
            self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面
            test_result_02 = pytest.assume(self.s20i_d(text="SIM card state(Searching)").exists(10))    # 判断potal界面是否一直显示查找SIM卡中
            if not test_result_02:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test09_step3_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test09 step3 change SIM 1 failed ScreenShot', allure.attachment_type.PNG)
        if all([test_result_01, test_result_02]):
            self.ws1['G10'] = "PASS"
        else:
            self.ws1['G10'] = "FAIL"

    def test11(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入实体卡后重启设备，选择云卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()
            self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_01 = pytest.assume(self.u3.login_type() == 1)
        with allure.step("第三步：插入另外一张SIM卡，登陆web界面切换SIM卡"):
            while True:
                inp = input('\033[35;0m%s' % '请插入另外一张实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
            self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_02 = pytest.assume(self.u3.login_type() == 0)  # 判断当前登陆类型是否为实体卡
            if not test_result_02:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test11_step3_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test11 step3 failed ScreenShot', allure.attachment_type.PNG)
        if all([test_result_01, test_result_02]):
            self.ws1['G12'] = "PASS"
        else:
            self.ws1['G12'] = "FAIL"

    def test12(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入实体卡后重启设备，选择云卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()
            self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_01 = pytest.assume(self.u3.login_type() == 1)
        with allure.step("第三步：重启设备，查看设备是否使用云卡登陆，且登陆成功"):
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_02 = pytest.assume(self.u3.login_type() == 1)
        if all([test_result_01, test_result_02]):
            self.ws1['G13'] = "PASS"
        else:
            self.ws1['G13'] = "FAIL"

    def test13(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：拔出实体卡后重启设备"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(1)
        with allure.step("第二步：插入实体卡后重启设备，选择云卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.open_webui()
            self.s20i_d(resourceId="text_sim_0").click()
            self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_01 = pytest.assume(self.u3.login_type() == 1)
        with allure.step("第三步：插入另外一张SIM卡，登陆web界面"):
            while True:
                inp = input('\033[35;0m%s' % '请插入另外一张实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.open_webui()
            test_result_02 = pytest.assume(self.s20i_d(text="A new physical SIM is detected, please select the way you want to use mobile data.").exists(10))
            if not test_result_02:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test013_step3_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test13 step3 failed ScreenShot', allure.attachment_type.PNG)
        if all([test_result_01, test_result_02]):
            self.ws1['G14'] = "PASS"
        else:
            self.ws1['G14'] = "FAIL"

    def test14(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：插入实体卡后重启设备，选择实体卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.open_webui()
            if self.s20i_d(text="SIM card selection").exists(5):
                self.s20i_d(resourceId="text_sim_1").click()
                self.s20i_d(text="Apply").click()
            else:
                self.login_webui()
                if self.s20i_d(text="In Use: Cloud SIM").exists(3) or self.s20i_d(text="In Use: Cloud SIM(Manual)").exists(3):
                    self.s20i_d(text="Settings").click()
                    self.s20i_d(text="SIM card management").click()
                    self.s20i_d(resourceId="sim_select").click()
                    self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
                    self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0    # 断言是否为实体卡登陆
        with allure.step("第二步：使用SIM卡过程中，登陆webui界面切换云卡"):
            time.sleep(60)  # 阻塞时间设置为60秒是为了防止云卡和实体卡频繁切换导致无法切换
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡登陆
            self.s20i_d(text="Apply").click()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result = pytest.assume(self.u3.login_type() == 1)  # 判断当前登陆类型是否为云卡
            if test_result:
                self.ws1['G15'] = "PASS"
            else:
                self.ws1['G15'] = "FAIL"

    def test15(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：插入实体卡后重启设备，选择实体卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)

                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.open_webui()
            if self.s20i_d(text="SIM card selection").exists(5):
                self.s20i_d(resourceId="text_sim_1").click()
                self.s20i_d(text="Apply").click()
            else:
                self.login_webui()
                if self.s20i_d(text="In Use: Cloud SIM").exists(3) or self.s20i_d(text="In Use: Cloud SIM(Manual)").exists(3):
                    self.s20i_d(text="Settings").click()
                    self.s20i_d(text="SIM card management").click()
                    self.s20i_d(resourceId="sim_select").click()
                    self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
                    self.s20i_d(text="Apply").click()
                    time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0    # 断言是否为实体卡登陆
        with allure.step("第二步：拔掉SIM卡，进入webui界面查看是否显示断开连接"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.open_webui()
            self.login_webui()
            # test_result_01 = pytest.assume(self.s20i_d(text="Status: Disconnect").exists(5))    # 判断网络是否断开连接
            test_result_01 = pytest.assume(self.s20i_d(text="Status: Absent").exists(5))    # 判断网络是否断开连接
            if not test_result_01:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test015_step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test15 step2 failed ScreenShot', allure.attachment_type.PNG)
        with allure.step("第三步：重启设备，查看云卡是否登陆成功"):
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_02 = pytest.assume(self.u3.login_type() == 1)  # 判断当前登陆类型是否为云卡
        if all([test_result_01, test_result_02]):
            self.ws1['G16'] = "PASS"
        else:
            self.ws1['G16'] = "FAIL"

    def test16(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：插入实体卡后重启设备，选择实体卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)

                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.open_webui()
            if self.s20i_d(text="SIM card selection").exists(5):
                self.s20i_d(resourceId="text_sim_1").click()
                self.s20i_d(text="Apply").click()
            else:
                self.login_webui()
                if self.s20i_d(text="In Use: Cloud SIM").exists(3) or self.s20i_d(text="In Use: Cloud SIM(Manual)").exists(3):
                    self.s20i_d(text="Settings").click()
                    self.s20i_d(text="SIM card management").click()
                    self.s20i_d(resourceId="sim_select").click()
                    self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
                    self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0    # 断言是否为实体卡登陆
        with allure.step("第二步：重启设备，查看是否使用SIM卡登陆"):
            self.u3.reboot()
            self.u3.wait_device_disconnect()
            self.u3.wait_device_connect()
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0  # 判断当前登陆类型是否为SIM卡
        with allure.step("第三步：进入webui界面切换云卡登陆"):
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result = pytest.assume(self.u3.login_type() == 1)  # 判断当前登陆类型是否为云卡
            if test_result:
                self.ws1['G17'] = "PASS"
            else:
                self.ws1['G17'] = "FAIL"

    def test17(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：插入实体卡后重启设备，选择实体卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)

                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.open_webui()
            if self.s20i_d(text="SIM card selection").exists(5):
                self.s20i_d(resourceId="text_sim_1").click()
                self.s20i_d(text="Apply").click()
            else:
                self.login_webui()
                if self.s20i_d(text="In Use: Cloud SIM").exists(3) or self.s20i_d(text="In Use: Cloud SIM(Manual)").exists(3):
                    self.s20i_d(text="Settings").click()
                    self.s20i_d(text="SIM card management").click()
                    self.s20i_d(resourceId="sim_select").click()
                    self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
                    self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0    # 断言是否为实体卡登陆
        with allure.step("第二步：拔掉SIM卡，进入webui界面查看是否显示断开连接"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.open_webui()
            self.login_webui()
            # test_result_01 = pytest.assume(self.s20i_d(text="Status: Disconnect").exists(5))    # 判断网络是否断开连接
            test_result_01 = pytest.assume(self.s20i_d(text="Status: Absent").exists(5))    # 判断网络是否断开连接
            if not test_result_01:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test017_step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test17 step2 failed ScreenShot', allure.attachment_type.PNG)
        with allure.step("第三步：插入SIM卡，查看设备网络是否连接成功"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result = pytest.assume(self.u3.login_type() == 0)  # 判断当前登陆类型是否为SIM卡
            if test_result:
                self.ws1['G18'] = "PASS"
            else:
                self.ws1['G18'] = "FAIL"

    def test18(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：插入实体卡后重启设备，选择实体卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)

                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.open_webui()
            if self.s20i_d(text="SIM card selection").exists(5):
                self.s20i_d(resourceId="text_sim_1").click()
                self.s20i_d(text="Apply").click()
            else:
                self.login_webui()
                if self.s20i_d(text="In Use: Cloud SIM").exists(3) or self.s20i_d(text="In Use: Cloud SIM(Manual)").exists(3):
                    self.s20i_d(text="Settings").click()
                    self.s20i_d(text="SIM card management").click()
                    self.s20i_d(resourceId="sim_select").click()
                    self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
                    self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0    # 断言是否为实体卡登陆
        with allure.step("第二步：拔掉SIM卡，进入webui界面查看是否显示断开连接"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.open_webui()
            self.login_webui()
            # test_result_01 = pytest.assume(self.s20i_d(text="Status: Disconnect").exists(5))    # 判断网络是否断开连接
            test_result_01 = pytest.assume(self.s20i_d(text="Status: Absent").exists(5))    # 判断网络是否断开连接
            if not test_result_01:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test018_step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test18 step2 failed ScreenShot', allure.attachment_type.PNG)
        with allure.step("第三步：切换云卡登陆"):
            if self.s20i_d(text="Wi-Fi clients").exists(3):
                pass
            else:
                self.open_webui()
                self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            time.sleep(3)
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 1  # 断言是否为云卡登陆
        with allure.step("第四步：切换SIM卡登陆"):
            time.sleep(60)
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
            self.s20i_d(text="Apply").click()
            time.sleep(30)
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
            self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
            self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面
            self.login_webui()
            test_result_02 = pytest.assume(self.s20i_d(text="Status: Connecting...").exists(5))  # 判断是否显示网络正在连接
            if not test_result_02:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test018_step4_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test18 step4 failed ScreenShot', allure.attachment_type.PNG)
        with allure.step("第五步：插入SIM卡，查看设备网络是否连接成功"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.wait_network_connect()
            time.sleep(3)
            test_result_03 = pytest.assume(self.u3.login_type() == 0)  # 判断当前登陆类型是否为SIM卡
        if all([test_result_01, test_result_02, test_result_03]):
            self.ws1['G19'] = "PASS"
        else:
            self.ws1['G19'] = "FAIL"

    def test19(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        time.sleep(0.5)
        with allure.step("第一步：插入实体卡后重启设备，选择实体卡登陆"):
            while True:
                inp = input('\033[35;0m%s' % '请插入实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)

                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.u3.reboot()
            time.sleep(5)
            self.open_webui()
            if self.s20i_d(text="SIM card selection").exists(5):
                self.s20i_d(resourceId="text_sim_1").click()
                self.s20i_d(text="Apply").click()
            else:
                self.login_webui()
                if self.s20i_d(text="In Use: Cloud SIM").exists(3) or self.s20i_d(text="In Use: Cloud SIM(Manual)").exists(3):
                    self.s20i_d(text="Settings").click()
                    self.s20i_d(text="SIM card management").click()
                    self.s20i_d(resourceId="sim_select").click()
                    self.s20i_d(resourceId="text_sim_1").click()  # 选择SIM卡
                    self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0    # 断言是否为实体卡登陆
        with allure.step("第二步：拔掉SIM卡，进入webui界面查看是否显示断开连接"):
            while True:
                inp = input('\033[35;0m%s' % '请拔出实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            self.open_webui()
            self.login_webui()
            test_result_01 = pytest.assume(self.s20i_d(text="Status: Disconnect").exists(5))    # 判断网络是否断开连接
            if not test_result_01:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test019_step2_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test19 step2 failed ScreenShot', allure.attachment_type.PNG)
        with allure.step("第三步：插入另外一张SIM卡，查看设备网络是否还是断开连接"):
            while True:
                inp = input('\033[35;0m%s' % '请插入另外一张实体卡, 确认操作完成后输入y继续(输入q退出测试)：')
                if inp == 'y':
                    break
                elif inp == 'q':
                    log.error('用户退出测试，结束测试, inp:%s' % inp)
                    assert False
                else:
                    log.error("输入错误，输入为：%s" % inp)
                    time.sleep(0.5)
                    continue
            time.sleep(30)
            self.open_webui()
            self.login_webui()
            test_result_02 = pytest.assume(self.s20i_d(text="Status: Disconnect").exists(5))  # 判断网络是否断开连接
            if not test_result_02:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test19_step3_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test19 step3 failed ScreenShot', allure.attachment_type.PNG)
        with allure.step("第四步：重启设备，查看webui界面是否显示选卡登陆界面"):
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.open_webui()
            test_result_03 = pytest.assume(self.s20i_d(text="A new physical SIM is detected, please select the way you want to use mobile data.").exists(10))
            if not test_result_03:
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test019_step4_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test19 step4 failed ScreenShot', allure.attachment_type.PNG)
        if all([test_result_01, test_result_02, test_result_03]):
            self.ws1['G20'] = "PASS"
        else:
            self.ws1['G20'] = "FAIL"

    def search_wifi(self, timeout=180):
        cur = time.time()
        expire = cur + timeout
        while cur < expire:
            if not self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(2):  # 判断当前界面是否存在名称为“GlocalMe_OHQPJE”的wifi
                while not self.s20i_d(resourceId="android:id/title", text="添加网络").exists(1):
                    if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(1):
                        break
                    else:
                        self.s20i_d.swipe(0.5, 0.8, 0.5, 0.3)
                else:
                    while not self.s20i_d(resourceId="android:id/title", text="WLAN 偏好设置").exists(1):
                        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(1):
                            break
                        else:
                            self.s20i_d.swipe(0.5, 0.3, 0.5, 0.8)
                cur = time.time()
            else:
                break
        else:
            raise TimeoutError("wifi connect timeout.")

    def connect_wifi(self):
        for i in range(2):
            self.s20i_d.press('home')
        self.s20i_d(resourceId="com.android.systemui:id/recent_apps").click()
        if self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").exists(3):
            self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").click()
            log.info("cleared background app success.")
        else:
            log.info('no background app running.')
        self.s20i_d.press('home')
        self.s20i_d(text="设置").click()
        self.s20i_d(resourceId="android:id/title", text="网络和互联网").click()
        self.s20i_d(resourceId="android:id/title", text="WLAN").click()
        if self.s20i_d(resourceId="com.android.settings:id/switch_widget").get_text() == "关闭":  # 判断wifi是否开启
            self.s20i_d(resourceId="com.android.settings:id/switch_widget").click()
        time.sleep(3)
        self.search_wifi()
        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
            return
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接")
            return
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
            return
        self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).click()  # 点击名称为“GlocalMe”的wifi
        if self.s20i_d(resourceId="com.android.settings:id/password").exists(3):  # 判断是否弹出输入password的弹窗
            self.s20i_d.send_keys(U3_settings.test_device_info['password'], clear=True)  # 输入password
            self.s20i_d(resourceId="android:id/button1").click()
        elif self.s20i_d(text="取消保存").exists(3):
            log.info("auxiliary device connect test device wifi successfully")
            return
        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接")
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
        else:
            log.error("auxiliary device connect test device wifi failed.")
            screenShot(self.s20i_d, title="wifi_connect_fail")
            raise ConnectionError("wifi connect error.")

    def open_webui(self):
        self.connect_wifi()
        self.s20i_d(resourceId="com.android.systemui:id/center_group").click()  # 返回主界面
        self.s20i_d(text="Chrome").click()  # 点击打开浏览器
        # 判断浏览器是否第一次打开，如果第一次打开就点击设置欢迎页弹窗
        if self.s20i_d(resourceId="com.android.chrome:id/terms_accept").exists(5):
            try:
                self.s20i_d(resourceId="com.android.chrome:id/terms_accept").click()
                self.s20i_d(resourceId="com.android.chrome:id/negative_button").click()
                self.s20i_d(resourceId="com.android.chrome:id/button_secondary").click()
            except Exception as e:
                log.error(e)
                # log.info('浏览器不是第一次启动，不需要设置欢迎页')
        # self.s20i_d(resourceId="com.android.chrome:id/home_button").click()  # 点击浏览器主页按钮进入主页
        try:
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
        except:
            self.s20i_d(resourceId="com.android.chrome:id/search_box").click()  # 点击浏览器搜索框
        self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
        self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面

    def login_webui(self):
        if self.s20i_d(resourceId="tr_manage_my_device").exists(5):
            self.s20i_d(resourceId="tr_manage_my_device").click()  # 点击管理我的设备
        elif self.s20i_d(text="Manage My Device").exists(5):
            self.s20i_d(text="Manage My Device").click()
        # 输入账号密码登陆管理界面
        self.s20i_d(resourceId="username").click()
        time.sleep(2)
        self.s20i_d.click(0.342, 0.303)
        self.s20i_d.send_keys("admin", clear=True)
        for i in range(2):
            self.s20i_d(resourceId="passWord").click()
        self.s20i_d.send_keys("admin", clear=True)
        self.s20i_d(text="Login").click()

    def teardown(self):
        # 每条case执行结束保存一下测试结果表
        self.wb.save(self.test_result_file_path)

    def teardown_class(self):
        log.info("-" * 20 + "End Test_CloudSimAndPhysicalSimSwitch" + "-" * 20)
        # self.d.open_notification()
        # self.d(resourceId="com.jy.recorder:id/tv_notify_stop").click()
        # self.start_SwitchComTool.kill()

@pytest.mark.skip
@allure.feature("Flash")
class Test_Flash:
    def setup_class(self):
        log.info("-" * 20 + "Start Test_Flash" + "-" * 20)
        self.u3 = models.U3(device_id=U3_settings.test_device_info['id'], log_project=log)
        # self.g4s = models.G4S(device_id=U3_settings.G4_info['id'], log_project=log)
        self.u3.wait_device_connect()
        # self.g4s.wait_device_connect()
        self._test_file = os.path.join(BASEDIR, 'config', 'test_file.zip')    # 测试文件路径
        self._test_storage_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'test_storage_path')    # 测试pull时文件的存放路径
        if not os.path.exists(self._test_storage_path):
            os.makedirs(self._test_storage_path)
        # self._test_file_size = os.path.getsize(self._test_file)
        self._test_file_size = round((os.path.getsize(self._test_file)/1000000), 2)    # 获取测试文件大小（字节），将单位转换为MB，并保留小数点2位
        log.info("测试文件大小：%s MB" % self._test_file_size)
        version = self.u3.get_current_version()  # 获取测试设备的版本
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)  # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path, '.'.join([version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', '整机测试内容_无屏MIFI.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["Flash"]  # 选择表

    def test01_push(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        u3_push_avg_result_list = []
        u3_push_duration_result_list = []
        # g4s_push_avg_result_list = []
        for i in range(50):
            log.info("-"*20 + "U3设备第%d次push" % (i+1) + "-"*20)
            u3_push_start_time = time.time()
            u3_push = self.u3.push(source_file=self._test_file, target_path='/sdcard/')
            u3_push_end_time = time.time()
            if u3_push:
                u3_push_duration = round((u3_push_end_time - u3_push_start_time), 2)    # 计算push时长
                u3_push_duration_result_list.append(u3_push_duration)
                log.info("push duration:%s" % u3_push_duration)
                u3_push_speed = round((self._test_file_size / u3_push_duration), 2)
                log.info("push speed:%s MB/S" % u3_push_speed)    # 计算push速率
                u3_push_avg_result_list.append(u3_push_speed)
            else:
                log.error("push faild.")
                u3_push_avg_result_list.append(0)
            time.sleep(1)
            # log.info("-" * 20 + "G4S设备第%d次push" % (i + 1) + "-" * 20)
            # g4s_start_time = time.time()
            # g4s_push = self.g4s.push(source_file=self._test_file, target_path='/sdcard/')    # 将测试文件push到设备sdcard目录下，返回push结果
            # g4s_end_time = time.time()
            # if g4s_push:
            #     g4s_push_duration = round((g4s_end_time - g4s_start_time), 2)  # 计算push时长
            #     log.info("push duration:%s" % g4s_push_duration)
            #     g4s_push_speed = round((self._test_file_size / g4s_push_duration), 2)
            #     log.info("push speed:%s MB/S" % g4s_push_speed)  # 计算push速率
            #     g4s_push_avg_result_list.append(g4s_push_speed)
            # else:
            #     log.error("push faild.")
            #     g4s_push_avg_result_list.append(0)
            # time.sleep(1)
        log.info('u3_push_avg_result_list:%s' % str(u3_push_avg_result_list))
        # log.info('g4s_push_avg_result_list:%s' % str(g4s_push_avg_result_list))
        u3_push_success_list = self._remove_value_from_list(u3_push_avg_result_list, 0)
        # g4s_push_success_list = self._remove_value_from_list(g4s_push_avg_result_list, 0)
        if len(u3_push_success_list) != len(u3_push_avg_result_list):
            u3_push_fail_times = len(u3_push_avg_result_list) - len(u3_push_success_list)
            log.error("u3 push fail times: %d" % u3_push_fail_times)
            self.ws1['H17'] = "u3 push fail times: %d" % u3_push_fail_times
        # if len(g4s_push_success_list) != len(g4s_push_avg_result_list):
        #     g4s_push_fail_times = len(g4s_push_avg_result_list) - len(g4s_push_success_list)
        #     log.error("g4s push fail times: %d" % g4s_push_fail_times)
        #     self.ws1['H18'] = "g4s push fail times: %d" % g4s_push_fail_times
        u3_push_avg = sum(u3_push_success_list) / len(u3_push_success_list)
        # g4s_push_avg = sum(g4s_push_success_list) / len(g4s_push_success_list)
        self.ws1['F17'] = round(u3_push_avg, 2)
        # self.ws1['F18'] = round(g4s_push_avg, 2)
        self.ws1['F32'] = len(u3_push_duration_result_list)
        self.ws1['G32'] = round((sum(u3_push_duration_result_list) / len(u3_push_duration_result_list)), 2)

    def test02_pull(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        u3_pull_result_list = []
        # g4s_pull_result_list = []
        u3_pull_duration_result_list = []
        # 判断U3设备/sdcard路径下是否存在test_file.zip文件，如果不存在就push一个进去
        if not self.u3.file_exists(file='/sdcard/test_file.zip'):
            u3_push = self.u3.push(source_file=self._test_file, target_path='/sdcard/')
            while not u3_push:
                u3_push = self.u3.push(source_file=self._test_file, target_path='/sdcard/')
        for i in range(50):
            log.info("-"*20 + "U3设备第%d次pull" % (i+1) + "-"*20)
            u3_pull_start_time = time.time()
            u3_pull = self.u3.pull(source_file="/sdcard/test_file.zip", target_path=self._test_storage_path)
            u3_pull_end_time = time.time()
            if u3_pull:
                u3_pull_duration = round((u3_pull_end_time - u3_pull_start_time), 2)    # 计算pull时长，保留小数点后2位
                u3_pull_duration_result_list.append(u3_pull_duration)
                log.info("pull duration:%s" % u3_pull_duration)
                u3_pull_speed = round((self._test_file_size / u3_pull_duration), 2)    # 计算pull速度，保留小数点后2位
                log.info("pull speed:%s" % u3_pull_speed)
                u3_pull_result_list.append(u3_pull_speed)
            else:
                log.error("pull failed.")
                u3_pull_result_list.append(0)
            time.sleep(1)
        log.info("u3_pull_result_list:%s" % str(u3_pull_result_list))
        u3_pull_success_list = self._remove_value_from_list(u3_pull_result_list, 0)    # 去除u3 pull结果集中为0的结果
        if len(u3_pull_success_list) != len(u3_pull_result_list):
            u3_pull_fail_times = len(u3_pull_result_list) - len(u3_pull_success_list)    # pull 失败次数
            log.error("u3 pull failed times: %d" % u3_pull_fail_times)
            self.ws1['H12'] = "u3 pull failed times: %d" % u3_pull_fail_times
        u3_pull_avg = round((sum(u3_pull_success_list) / len(u3_pull_success_list)), 2)
        self.ws1['F12'] = u3_pull_avg
        self.ws1['F31'] = len(u3_pull_duration_result_list)
        self.ws1['G31'] = round((sum(u3_pull_duration_result_list) / len(u3_pull_duration_result_list)), 2)
        '''
        # 判断G4S设备/sdcard路径下是否存在test_file.zip文件，如果不存在就push一个进去
        if not self.g4s.file_exists(file='/sdcard/test_file.zip'):
            g4s_push = self.g4s.push(source_file=self._test_file, target_path='/sdcard/')
            while not g4s_push:
                g4s_push = self.g4s.push(source_file=self._test_file, target_path='/sdcard/')
        for i in range(50):
            log.info("-" * 20 + "G4S设备第%d次pull" % (i + 1) + "-" * 20)
            g4s_pull_start_time = time.time()
            g4s_pull = self.g4s.pull(source_file="/sdcard/test_file.zip", target_path=self._test_storage_path)
            g4s_pull_end_time = time.time()
            if g4s_pull:
                g4s_pull_duration = round((g4s_pull_end_time - g4s_pull_start_time), 2)  # 计算pull时长，保留小数点后2位
                log.info("pull duration:%s" % g4s_pull_duration)
                g4s_pull_speed = round((self._test_file_size / g4s_pull_duration), 2)  # 计算pull速度，保留小数点后2位
                log.info("pull speed:%s" % g4s_pull_speed)
                g4s_pull_result_list.append(g4s_pull_speed)
            else:
                log.error("pull failed.")
                g4s_pull_result_list.append(0)
            time.sleep(1)
        log.info("g4s_pull_result_list:%s" % str(g4s_pull_result_list))
        g4s_pull_success_list = self._remove_value_from_list(g4s_pull_result_list, 0)  # 去除g4s pull结果集中为0的结果
        if len(g4s_pull_success_list) != len(g4s_pull_result_list):
            g4s_pull_fail_times = len(g4s_pull_result_list) - len(g4s_pull_success_list)  # pull 失败次数
            log.error("g4s pull failed times: %d" % g4s_pull_fail_times)
            self.ws1['H13'] = "g4s pull failed times: %d" % g4s_pull_fail_times
        g4s_pull_avg = round((sum(g4s_pull_success_list) / len(g4s_pull_success_list)), 2)    # 计算pull成功结果的平均值，保留小数点2位
        self.ws1['F13'] = g4s_pull_avg
        '''

    def teardown(self):
        # 每条case执行结束保存一下测试结果表
        self.wb.save(self.test_result_file_path)

    def teardown_class(self):
        log.info("-" * 20 + "End Test_Flash" + "-" * 20)

    def _remove_value_from_list(self, list, value):
        for i in list:
            if i == value:
                list.remove(i)
        return list

@pytest.mark.skip
@allure.feature("性能指标")
class Test_PerformanceIndex:
    def setup_class(self):
        log.info("-" * 20 + "Start Test_PerformanceIndex" + "-" * 20)
        self.u3 = models.U3(device_id=U3_settings.test_device_info['id'], log_project=log)
        self.u3.wait_device_connect()
        version = self.u3.get_current_version()  # 获取测试设备的版本
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)  # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path, '.'.join([version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', '整机测试内容_无屏MIFI.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["性能指标"]  # 选择表

    def test01(self):
        '''测试用例：重启'''
        pass

    def teardown(self):
        # 每条case执行结束保存一下测试结果表
        self.wb.save(self.test_result_file_path)

    def teardown_class(self):
        log.info("-" * 20 + "End Test_PerformanceIndex" + "-" * 20)

@allure.feature("显示")
class Test_Display:
    def setup_class(self):
        log.info("-" * 20 + "Start Test_Display" + "-" * 20)
        self.u3 = models.U3(device_id=U3_settings.test_device_info['id'], log_project=log)
        self.s20i = models.Glocalme(device_id=U3_settings.auxiliary_device_info['id'], log_project=log)
        self.u3.wait_device_connect()
        self.s20i.wait_device_connect()
        self.version = self.u3.get_current_version()  # 获取测试设备的版本
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)  # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path, '.'.join([self.version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', '整机测试内容_无屏MIFI.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["显示"]  # 选择表
        # 执行uiautomator2的init操作，在测试设备端安装uiautomator app，minicap和minitouch，atx-agent
        uiautomator2_init = subprocess.getoutput("python -m uiautomator2 init")
        self.s20i_d = u2.connect('c58790e3')
        self.s20i.wakey()  # 设置辅助机屏幕常亮
        # 判断S20i屏幕是否锁屏状态，如果锁屏就给设备解锁（仅适用滑动解锁）
        if self.s20i_d(resourceId="com.android.systemui:id/lock_icon").exists():
            self.s20i_d.xpath('//android.widget.FrameLayout').click()
            self.s20i_d(resourceId="com.android.systemui:id/lock_icon").click()
        # '''查找并连接wifi'''
        # for i in range(2):
        #     self.s20i_d.press('home')
        # # 杀掉后台程序
        # self.s20i_d(resourceId="com.android.systemui:id/recent_apps").click()
        # if self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").exists(3):
        #     self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").click()
        #     log.info("cleared background app success.")
        # else:
        #     log.info('no background app running.')
        # '''进入wifi设置界面查找wifi并连接wifi'''
        # self.s20i_d.press('home')
        # self.s20i_d(text="设置").click()
        # self.s20i_d(resourceId="android:id/title", text="网络和互联网").click()
        # self.s20i_d(resourceId="android:id/title", text="WLAN").click()
        # if self.s20i_d(resourceId="com.android.settings:id/switch_widget").get_text() == "关闭":  # 判断wifi是否开启
        #     self.s20i_d(resourceId="com.android.settings:id/switch_widget").click()
        # time.sleep(3)
        # cur = time.time()
        # expire = cur + 180
        # while cur < expire:
        #     if not self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(2):  # 判断当前界面是否存在名称为“GlocalMe_OHQPJE”的wifi
        #         while not self.s20i_d(resourceId="android:id/title", text="添加网络").exists(1):
        #             if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(1):
        #                 break
        #             else:
        #                 self.s20i_d.swipe(0.5, 0.8, 0.5, 0.3)
        #         else:
        #             while not self.s20i_d(resourceId="android:id/title", text="WLAN 偏好设置").exists(1):
        #                 if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(
        #                         1):
        #                     break
        #                 else:
        #                     self.s20i_d.swipe(0.5, 0.3, 0.5, 0.8)
        #         cur = time.time()
        #     else:
        #         break
        # else:
        #     raise TimeoutError("wifi connect timeout.")
        # if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
        #     log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
        #     return
        # elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
        #     log.info("auxiliary device connect test device wifi successfully, text:已连接")
        #     return
        # elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
        #     log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
        #     return
        # self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).click()  # 点击名称为“GlocalMe”的wifi
        # if self.s20i_d(resourceId="com.android.settings:id/password").exists(3):  # 判断是否弹出输入password的弹窗
        #     self.s20i_d.send_keys(U3_settings.test_device_info['password'], clear=True)  # 输入password
        #     self.s20i_d(resourceId="android:id/button1").click()
        # elif self.s20i_d(text="取消保存").exists(3):
        #     log.info("auxiliary device connect test device wifi successfully")
        #     return
        # if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
        #     log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
        # elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
        #     log.info("auxiliary device connect test device wifi successfully, text:已连接")
        # elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
        #     log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
        # else:
        #     log.error("auxiliary device connect test device wifi failed.")
        #     screenShot(self.s20i_d, title="wifi_connect_fail")
        #     raise ConnectionError("wifi connect error.")
        # '''进入webui首页'''
        # self.s20i_d(resourceId="com.android.systemui:id/center_group").click()  # 返回主界面
        # self.s20i_d(text="Chrome").click()  # 点击打开浏览器
        # # 判断浏览器是否第一次打开，如果第一次打开就点击设置欢迎页弹窗
        # if self.s20i_d(resourceId="com.android.chrome:id/terms_accept").exists(5):
        #     try:
        #         self.s20i_d(resourceId="com.android.chrome:id/terms_accept").click()
        #         self.s20i_d(resourceId="com.android.chrome:id/negative_button").click()
        #         self.s20i_d(resourceId="com.android.chrome:id/button_secondary").click()
        #     except Exception as e:
        #         log.error(e)
        #         # log.info('浏览器不是第一次启动，不需要设置欢迎页')
        # # self.s20i_d(resourceId="com.android.chrome:id/home_button").click()  # 点击浏览器主页按钮进入主页
        # try:
        #     self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
        # except:
        #     self.s20i_d(resourceId="com.android.chrome:id/search_box").click()  # 点击浏览器搜索框
        # self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
        # self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面
        # time.sleep(3)


    def test01(self):
        '''测试用例：卡状态'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        card_info_list = [
            "In Use: Cloud SIM",
            "In Use: Cloud SIM(Manual)",
            "In Use: SIM 1",
            "In Use: SIM 1(Manual)",
        ]
        status_info_list = [
            "Status: Connecting...",
            "Status: Connected",
            "Status: Disconnect",
        ]
        self.open_webui()
        time.sleep(3)
        card_info = self.s20i_d(resourceId="div_manage_my_device").child().child()[0].child().get_text()    # 获取In Use 信息
        log.info("card_info:%s" % card_info)
        status_info = self.s20i_d(resourceId="div_manage_my_device").child().child()[1].child().get_text()    # 获取Status 信息
        log.info("status_info:%s" % status_info)
        test_result_01 = pytest.assume(card_info in card_info_list)
        test_result_02 = pytest.assume(status_info in status_info_list)
        test_result_03 = pytest.assume(self.s20i_d(text="Manage My Device").exists(3))
        if all([test_result_01, test_result_02, test_result_03]):
            self.ws1['G8'] = "PASS"
        else:
            self.ws1['G8'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test01_CardStatus_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test01_CardStatus failed ScreenShot', allure.attachment_type.PNG)

    def test02(self):
        '''测试用例：WiFi信息及版本'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        display_name = [
            "Wi-Fi SSID:",
            "Client Num:",
            "Version:",
        ]
        self.open_webui()
        time.sleep(3)
        name01 = self.s20i_d(resourceId="status_info").child()[12].get_text()    # 获取ssid显示的名称
        name02 = self.s20i_d(resourceId="status_info").child()[14].get_text()    # 获取客户端数量显示的名称
        name03 = self.s20i_d(resourceId="status_info").child()[16].get_text()    # 获取版本信息显示的名称
        log.info("ssid显示的名称：%s, 客户端数量显示的名称：%s, 版本信息显示的名称：%s" % (name01, name02, name03))
        ssid = self.s20i_d(resourceId="status_info").child()[13].get_text()    # 获取ssid
        client_num = self.s20i_d(resourceId="status_info").child()[15].get_text()    # 获取客户端连接数
        version = self.s20i_d(resourceId="status_info").child()[17].get_text()    # 获取版本号
        log.info("ssid:%s, client_num:%s, version:%s" % (ssid, client_num, version))
        test_result_01 = pytest.assume(name01 == display_name[0])
        test_result_02 = pytest.assume(name02 == display_name[1])
        test_result_03 = pytest.assume(name03 == display_name[2])
        test_result_04 = pytest.assume(ssid == U3_settings.test_device_info['ssid'])
        test_result_05 = pytest.assume(int(client_num) > 0)
        test_result_06 = pytest.assume(version[0:-4] in self.version)
        if all([test_result_01, test_result_02, test_result_03, test_result_04, test_result_05, test_result_06]):
            self.ws1['G9'] = "PASS"
        else:
            self.ws1['G9'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test02_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test02 test failed ScreenShot', allure.attachment_type.PNG)

    def test03(self):
        '''测试用例：WiFi名称'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        time.sleep(3)
        ssid = self.s20i_d(resourceId="ssid").get_text()
        log.info("Wi-Fi SSID:%s" % ssid)
        test_result = pytest.assume(ssid = U3_settings.test_device_info['ssid'])
        if test_result:
            self.ws1['G10'] = "PASS"
        else:
            self.ws1['G10'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test03_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test03 test failed ScreenShot', allure.attachment_type.PNG)

    def test04(self):
        '''测试用例：WiFi密码'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        time.sleep(3)
        password = self.s20i_d(resourceId="password").get_text()
        log.info("Password:%s" % password)
        test_result = pytest.assume(password = U3_settings.test_device_info['password'])
        if test_result:
            self.ws1['G11'] = "PASS"
        else:
            self.ws1['G11'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test04_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test04 test failed ScreenShot', allure.attachment_type.PNG)

    def test05(self):
        '''测试用例：套餐显示以及卡状态'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        card_info_list = [
            "In Use: Cloud SIM",
            "In Use: Cloud SIM(Manual)",
            "In Use: SIM 1",
            "In Use: SIM 1(Manual)",
        ]
        status_info_list = [
            "Status: Connecting...",
            "Status: Connected",
            "Status: Disconnect",
        ]
        self.open_webui()
        self.login_webui()
        time.sleep(3)
        card_info = self.s20i_d(resourceId="div_wifi_device_info").child().child()[0].child().get_text()
        status_info = self.s20i_d(resourceId="div_wifi_device_info").child().child()[1].child().get_text()
        log.info("card_info:%s, status_info:%s" % (card_info, status_info))
        test_result_01 = pytest.assume(card_info in card_info_list)
        test_result_02 = pytest.assume(status_info in status_info_list)
        if all([test_result_01, test_result_02]):
            self.ws1['G12'] = "PASS"
        else:
            self.ws1['G12'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test05_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test05 test failed ScreenShot', allure.attachment_type.PNG)

    def test06(self):
        '''测试用例：显示连接设备数'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Wi-Fi clients").click()
        time.sleep(3)
        client_connect_info = self.s20i_d(resourceId="div_list_info").child().get_text()    # 获取UI界面显示的客户端连接信息
        client_num = self.s20i_d(resourceId="ul_list").info['childCount']    # 获取连接的设备数量
        expire_connect_info = ' '.join([str(client_num), 'devices are securely connected to Internet...'])    # 预期UI界面显示的客户端连接信息
        test_result = pytest.assume(client_connect_info == expire_connect_info)
        if test_result:
            self.ws1['G13'] = "PASS"
        else:
            self.ws1['G13'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test06_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test06 test failed ScreenShot', allure.attachment_type.PNG)

    def test07(self):
        '''测试用例：显示连接设备品牌'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Wi-Fi clients").click()
        time.sleep(3)
        brand_list = [
            "Other Devices",
            "Apple Devices",
        ]
        test_result_list = []    # 创建一个测试结果集
        acquired_brand_list = []    # 创建一个获取到的品牌集
        child_object = self.s20i_d(resourceId="ul_list").child()    # 获取resourceId="ul_list"的对象的所有子对象
        # print(len(child_object))
        client_num = int(len(child_object) / 6)    # 根据获取到的子对象来计算连接的客户端数量
        for i in range(client_num):
            brand = self.s20i_d(resourceId="ul_list").child()[2 + 6 * i].get_text()    # 获取每个客户端的设备品牌
            acquired_brand_list.append(brand)    # 将获取到的每个客户端的设备品牌加入到acquired_brand_list
            test_result = pytest.assume(brand in brand_list)    # 判断获取的每个客户端的设备品牌是否在品牌集中
            test_result_list.append(test_result)    # 将测试结果加入测试结果集
        log.info("acquired_brand_list:%s" % str(acquired_brand_list))
        log.info("test_result_list:%s" % str(test_result_list))
        if all(test_result_list):
            self.ws1['G14'] = "PASS"
        else:
            self.ws1['G14'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test07_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test07 test failed ScreenShot', allure.attachment_type.PNG)

    def test08(self):
        '''测试用例：显示黑名单设备'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Wi-Fi clients").click()
        time.sleep(3)
        if self.s20i_d(resourceId="div_wifi_black_lists").exists(3):
            self.s20i_d(resourceId="div_wifi_black_lists").click()
            test_result = pytest.assume(self.s20i_d(resourceId="ul_list").exists(3))
            if test_result:
                self.ws1['G15'] = "PASS"
            else:
                self.ws1['G15'] = "FAIL"
                FailScreenShot_file = screenShot(d=self.s20i_d, title='test08_fail')
                file = open(FailScreenShot_file, 'rb').read()
                allure.attach(file, 'test08 test failed ScreenShot', allure.attachment_type.PNG)
        else:
            self.ws1['G15'] = "NA"
            self.ws1['H15'] = "UI界面没有显示黑名单列表"
            assert False

    def test09(self):
        '''测试用例：设备加入黑名单开关'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Wi-Fi clients").click()
        time.sleep(3)
        test_result_list = []  # 创建一个测试结果集
        client_num = self.s20i_d(resourceId="ul_list").info['childCount']  # 获取连接的设备数量
        for i in range(client_num):
            resourceId = '_'.join(['wifi_client_lists', str(i)])    # 每个客户端加入黑名单开关的resourceId
            test_result = pytest.assume(self.s20i_d(resourceId=resourceId).exists(3))    # 判断每个客户端加入黑名单的开关是否存在
            test_result_list.append(test_result)
        log.info("test_result_list:%s" % str(test_result_list))
        if all(test_result_list):
            self.ws1['G16'] = "PASS"
        else:
            self.ws1['G16'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test09_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test09 test failed ScreenShot', allure.attachment_type.PNG)

    def test10(self):
        '''测试用例：连接设备的MAC地址'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Wi-Fi clients").click()
        time.sleep(3)
        test_result_list = []  # 创建一个测试结果集
        acquired_mac_list = []  # 创建一个获取到的品牌集
        child_object = self.s20i_d(resourceId="ul_list").child()  # 获取resourceId="ul_list"的对象的所有子对象
        # print(len(child_object))
        client_num = int(len(child_object) / 6)  # 根据获取到的子对象来计算连接的客户端数量
        for i in range(client_num):
            mac = self.s20i_d(resourceId="ul_list").child()[3 + 6 * i].get_text()  # 获取每个客户端的设备MAC地址
            acquired_mac_list.append(mac)  # 将获取到的每个客户端的设备MAC地址加入到acquired_mac_list
            test_result = pytest.assume(len(mac) == 17)  # 判断获取的每个客户端的MAC地址字符串长度是否为17
            test_result_list.append(test_result)  # 将测试结果加入测试结果集
        log.info("acquired_mac_list:%s" % str(acquired_mac_list))
        log.info("test_result_list:%s" % str(test_result_list))
        if all(test_result_list):
            self.ws1['G17'] = "PASS"
        else:
            self.ws1['G17'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test10_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test10 test failed ScreenShot', allure.attachment_type.PNG)

    def test11(self):
        '''测试用例：设置界面UI显示'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Settings").click()
        time.sleep(3)
        test_result_list = []
        test_result_01 = pytest.assume(self.s20i_d(text="Web administrator").exists(3))    # 判断设置界面是否显示UI “Web administrator”
        test_result_list.append(test_result_01)
        if test_result_01:
            self.ws1['G18'] = "PASS"
        else:
            self.ws1['G18'] = "FAIL"
        test_result_02 = pytest.assume(self.s20i_d(text="SIM card management").exists(3))    # 判断设置界面是否显示UI “SIM card management”
        test_result_list.append(test_result_02)
        if test_result_02:
            self.ws1['G19'] = "PASS"
        else:
            self.ws1['G19'] = "FAIL"
        test_result_03 = pytest.assume(self.s20i_d(text="Data management").exists(3))  # 判断设置界面是否显示UI “Data management”
        test_result_list.append(test_result_03)
        if test_result_03:
            self.ws1['G20'] = "PASS"
        else:
            self.ws1['G20'] = "FAIL"
        test_result_04 = pytest.assume(self.s20i_d(text="Language settings").exists(3) and self.s20i_d(text="English").exists(3))  # 判断设置界面是否显示UI “Language settings”和“English”
        test_result_list.append(test_result_04)
        if test_result_04:
            self.ws1['G21'] = "PASS"
        else:
            self.ws1['G21'] = "FAIL"
        test_result_05 = pytest.assume(self.s20i_d(text="Restore to factory defaults").exists(3))  # 判断设置界面是否显示UI “Restore to factory defaults”
        test_result_list.append(test_result_05)
        if test_result_05:
            self.ws1['G22'] = "PASS"
        else:
            self.ws1['G22'] = "FAIL"
        test_result_06 = pytest.assume(self.s20i_d(text="Logout").exists(3))  # 判断设置界面是否显示UI “Logout”
        test_result_list.append(test_result_06)
        if test_result_06:
            self.ws1['G23'] = "PASS"
        else:
            self.ws1['G23'] = "FAIL"
        test_result_07 = pytest.assume(self.s20i_d(text="About").exists(3))  # 判断设置界面是否显示UI “About”
        test_result_list.append(test_result_07)
        if test_result_07:
            self.ws1['G24'] = "PASS"
        else:
            self.ws1['G24'] = "FAIL"
        log.info("test_result_list:%s" % str(test_result_list))
        if not all(test_result_list):
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test11_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test11 test failed ScreenShot', allure.attachment_type.PNG)

    def teardown(self):
        # 每条case执行结束保存一下测试结果表
        self.wb.save(self.test_result_file_path)

    def teardown_class(self):
        log.info("-" * 20 + "End Test_Display" + "-" * 20)

    def search_wifi(self, timeout=180):
        cur = time.time()
        expire = cur + timeout
        while cur < expire:
            if not self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(2):  # 判断当前界面是否存在名称为“GlocalMe_OHQPJE”的wifi
                while not self.s20i_d(resourceId="android:id/title", text="添加网络").exists(1):
                    if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(1):
                        break
                    else:
                        self.s20i_d.swipe(0.5, 0.8, 0.5, 0.3)
                else:
                    while not self.s20i_d(resourceId="android:id/title", text="WLAN 偏好设置").exists(1):
                        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(1):
                            break
                        else:
                            self.s20i_d.swipe(0.5, 0.3, 0.5, 0.8)
                cur = time.time()
            else:
                break
        else:
            raise TimeoutError("wifi connect timeout.")

    def connect_wifi(self):
        for i in range(2):
            self.s20i_d.press('home')
        self.s20i_d(resourceId="com.android.systemui:id/recent_apps").click()
        if self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").exists(3):
            self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").click()
            log.info("cleared background app success.")
        else:
            log.info('no background app running.')
        self.s20i_d.press('home')
        self.s20i_d(text="设置").click()
        self.s20i_d(resourceId="android:id/title", text="网络和互联网").click()
        self.s20i_d(resourceId="android:id/title", text="WLAN").click()
        if self.s20i_d(resourceId="com.android.settings:id/switch_widget").get_text() == "关闭":  # 判断wifi是否开启
            self.s20i_d(resourceId="com.android.settings:id/switch_widget").click()
        time.sleep(3)
        self.search_wifi()
        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
            return
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接")
            return
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
            return
        self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).click()  # 点击名称为“GlocalMe”的wifi
        if self.s20i_d(resourceId="com.android.settings:id/password").exists(3):  # 判断是否弹出输入password的弹窗
            self.s20i_d.send_keys(U3_settings.test_device_info['password'], clear=True)  # 输入password
            self.s20i_d(resourceId="android:id/button1").click()
        elif self.s20i_d(text="取消保存").exists(3):
            log.info("auxiliary device connect test device wifi successfully")
            return
        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接")
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
        else:
            log.error("auxiliary device connect test device wifi failed.")
            screenShot(self.s20i_d, title="wifi_connect_fail")
            raise ConnectionError("wifi connect error.")

    def open_webui(self):
        self.connect_wifi()
        self.s20i_d(resourceId="com.android.systemui:id/center_group").click()  # 返回主界面
        self.s20i_d(text="Chrome").click()  # 点击打开浏览器
        # 判断浏览器是否第一次打开，如果第一次打开就点击设置欢迎页弹窗
        if self.s20i_d(resourceId="com.android.chrome:id/terms_accept").exists(5):
            try:
                self.s20i_d(resourceId="com.android.chrome:id/terms_accept").click()
                self.s20i_d(resourceId="com.android.chrome:id/negative_button").click()
                self.s20i_d(resourceId="com.android.chrome:id/button_secondary").click()
            except Exception as e:
                log.error(e)
                # log.info('浏览器不是第一次启动，不需要设置欢迎页')
        # self.s20i_d(resourceId="com.android.chrome:id/home_button").click()  # 点击浏览器主页按钮进入主页
        try:
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
        except:
            self.s20i_d(resourceId="com.android.chrome:id/search_box").click()  # 点击浏览器搜索框
        self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
        self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面

    def login_webui(self):
        if self.s20i_d(resourceId="tr_manage_my_device").exists(5):
            self.s20i_d(resourceId="tr_manage_my_device").click()  # 点击管理我的设备
        elif self.s20i_d(text="Manage My Device").exists(5):
            self.s20i_d(text="Manage My Device").click()
        # 输入账号密码登陆管理界面
        self.s20i_d(resourceId="username").click()
        time.sleep(2)
        self.s20i_d.click(0.342, 0.303)
        self.s20i_d.send_keys("admin", clear=True)
        for i in range(2):
            self.s20i_d(resourceId="passWord").click()
        self.s20i_d.send_keys("admin", clear=True)
        self.s20i_d(text="Login").click()

@allure.feature("加入云卡后的稳定性")
class Test_SystemStability:
    def setup_class(self):
        log.info("-" * 20 + "Start Test_SystemStability" + "-" * 20)
        self.u3 = models.U3(device_id=U3_settings.test_device_info['id'], log_project=log)
        self.s20i = models.Glocalme(device_id=U3_settings.auxiliary_device_info['id'], log_project=log)
        self.u3.wait_device_connect()
        self.s20i.wait_device_connect()
        self.version = self.u3.get_current_version()  # 获取测试设备的版本
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)  # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path,'.'.join([self.version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', '整机测试内容_无屏MIFI.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["加入云卡后的稳定性"]  # 选择表
        # 执行uiautomator2的init操作，在测试设备端安装uiautomator app，minicap和minitouch，atx-agent
        uiautomator2_init = subprocess.getoutput("python -m uiautomator2 init")
        self.s20i_d = u2.connect('c58790e3')
        self.s20i.wakey()  # 设置辅助机屏幕常亮
        # 判断S20i屏幕是否锁屏状态，如果锁屏就给设备解锁（仅适用滑动解锁）
        if self.s20i_d(resourceId="com.android.systemui:id/lock_icon").exists():
            self.s20i_d.xpath('//android.widget.FrameLayout').click()
            self.s20i_d(resourceId="com.android.systemui:id/lock_icon").click()

    @pytest.mark.skip
    def test01(self):
        '''测试用例：开关机(云卡)100次'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        login_type = self.u3.login_type()
        if login_type != 1:
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type()==1
        test_success_times = 0
        for i in range(100):
            log.info("-" * 20 + "第%d次重启测试" % (i + 1) + "-" * 20)
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            self.u3.wait_network_connect()
            time.sleep(3)
            login_type_01 = self.u3.login_type()
            test_result = pytest.assume(login_type_01==1)
            if test_result:
                test_success_times += 1
            else:
                self.ws1['K2'] = "FAIL"
                self.ws1['L2'] = "重启测试成功次数：%d" % test_success_times
                log.error("test_result:%s, u3 login type:%s, test success times:%d" % (test_result, login_type_01, test_success_times))
                assert False
        self.ws1['K2'] = "PASS"

    def test09(self):
        '''测试用例：关闭云卡后实体卡上网'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        login_type = self.u3.login_type()
        if login_type != 1:
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 1
            time.sleep(30)
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Settings").click()
        self.s20i_d(text="SIM card management").click()
        self.s20i_d(resourceId="sim_select").click()
        self.s20i_d(resourceId="text_sim_1").click()  # 选择实体卡卡
        self.s20i_d(text="Apply").click()
        time.sleep(5)
        self.u3.wait_network_connect()
        time.sleep(3)
        login_type_01 = self.u3.login_type()
        test_result = pytest.assume(login_type_01 == 0)
        if test_result:
            self.ws1['K10'] = "PASS"
        else:
            self.ws1['k10'] = "FAIL"
            log.error('U3 login type:%s' % login_type_01)

    @pytest.mark.skip
    def test11(self):
        '''测试用例：启动云卡过程关机再开机'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.u3.wait_network_connect()
        login_type = self.u3.login_type()
        if login_type != 1:
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 1
        test_success_times = 0
        for i in range(100):
            log.info("-" * 20 + "第%d次启动云卡过程中重启测试" % (i + 1) + "-" * 20)
            self.u3.reboot()
            time.sleep(5)
            self.u3.wait_device_connect()
            sleep_duration = random.randint(10,40)    # 生成一个10-40之间的随机数，将其设置为设备重启后等待时长（比拨号成功时间短）
            log.info("sleep duration:%d" % sleep_duration)
            time.sleep(sleep_duration)
            if self.u3.connect_network():
                self.ws1['K12'] = "FAIL"
                self.ws1['L12'] = "启动云卡过程中重启测试成功次数：%d" % test_success_times
                assert False
            test_success_times += 1
        self.ws1['K12'] = "PASS"

    @pytest.mark.skip
    def test14(self):
        '''测试用例：云卡ping包2小时'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.u3.wait_network_connect()
        login_type = self.u3.login_type()
        if login_type != 1:
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
            self.s20i_d(text="Apply").click()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 1
        ping_result = self.u3.ping(duration=7200, target="www.baidu.com")     # ping "www.baidu.com" 7200秒
        packet_loss = ping_result[2]    # 丢包率
        if packet_loss > 5:
            self.ws1['K15'] = "FAIL"
            self.ws1['L15'] = "%d packets transmitted, %d received, %.2f%% packet loss" % ping_result
            assert False
        else:
            self.ws1['K15'] = "PASS"

    def test16(self):
        '''测试用例：云卡网络设置'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.u3.wait_network_connect()
        login_type = self.u3.login_type()
        if login_type != 0:
            self.open_webui()
            self.login_webui()
            self.s20i_d(text="Settings").click()
            self.s20i_d(text="SIM card management").click()
            self.s20i_d(resourceId="sim_select").click()
            self.s20i_d(resourceId="text_sim_1").click()  # 选择实体卡
            self.s20i_d(text="Apply").click()
            time.sleep(5)
            self.u3.wait_network_connect()
            time.sleep(3)
            assert self.u3.login_type() == 0
        self.open_webui()
        self.login_webui()
        self.s20i_d(text="Settings").click()
        self.s20i_d(text="SIM card management").click()
        # sim_roam_dict = self.s20i_d(resourceId="sim_roam").info
        data_roaming = self.s20i_d(resourceId="sim_roam").info['checked']    # 获取漫游开关是否打开，打开为True，没打开为False
        # 打开漫游开关
        if not data_roaming:
            self.s20i_d(resourceId="sim_roam").click()
            if self.s20i_d(text="Are you sure to enable data roaming for this physical SIM card?").exists(10):
                self.s20i_d(resourceId="confirm_btn_ok").click()
            time.sleep(3)
        # data_roaming_on = self.s20i_d(resourceId="sim_roam").info['checked']
        # assert data_roaming_on == True
        try:
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
        except:
            self.s20i_d(resourceId="com.android.chrome:id/search_box").click()  # 点击浏览器搜索框
        self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
        self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面
        self.login_webui()
        self.s20i_d(text="Settings").click()
        self.s20i_d(text="SIM card management").click()
        self.s20i_d(resourceId="sim_select").click()
        self.s20i_d(resourceId="text_sim_0").click()  # 选择云卡
        self.s20i_d(text="Apply").click()
        time.sleep(5)
        self.u3.wait_network_connect()
        time.sleep(3)
        assert self.u3.login_type() == 1
        time.sleep(30)
        try:
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
        except:
            self.s20i_d(resourceId="com.android.chrome:id/search_box").click()  # 点击浏览器搜索框
        self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
        self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面
        self.login_webui()
        self.s20i_d(text="Settings").click()
        self.s20i_d(text="SIM card management").click()
        self.s20i_d(resourceId="sim_select").click()
        self.s20i_d(resourceId="text_sim_1").click()  # 选择实体卡
        self.s20i_d(text="Apply").click()
        time.sleep(5)
        self.u3.wait_network_connect()
        time.sleep(3)
        assert self.u3.login_type() == 0
        try:
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
        except:
            self.s20i_d(resourceId="com.android.chrome:id/search_box").click()  # 点击浏览器搜索框
        self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
        self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面
        self.login_webui()
        self.s20i_d(text="Settings").click()
        self.s20i_d(text="SIM card management").click()
        # sim_roam_dict = self.s20i_d(resourceId="sim_roam").info
        data_roaming_on = self.s20i_d(resourceId="sim_roam").info['checked']  # 获取漫游开关是否打开，打开为True，没打开为False
        test_result = pytest.assume(data_roaming_on == True)
        if test_result:
            self.ws1['K17'] = "PASS"
        else:
            self.ws1['K17'] = "FAIL"
            FailScreenShot_file = screenShot(d=self.s20i_d, title='test16_fail')
            file = open(FailScreenShot_file, 'rb').read()
            allure.attach(file, 'test16 test failed ScreenShot', allure.attachment_type.PNG)


    def teardown(self):
        # 每条case执行结束保存一下测试结果表
        self.wb.save(self.test_result_file_path)

    def teardown_class(self):
        log.info("-" * 20 + "End Test_SystemStability" + "-" * 20)

    def search_wifi(self, timeout=180):
        cur = time.time()
        expire = cur + timeout
        while cur < expire:
            if not self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(2):  # 判断当前界面是否存在名称为“GlocalMe_OHQPJE”的wifi
                while not self.s20i_d(resourceId="android:id/title", text="添加网络").exists(1):
                    if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(1):
                        break
                    else:
                        self.s20i_d.swipe(0.5, 0.8, 0.5, 0.3)
                else:
                    while not self.s20i_d(resourceId="android:id/title", text="WLAN 偏好设置").exists(1):
                        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).exists(1):
                            break
                        else:
                            self.s20i_d.swipe(0.5, 0.3, 0.5, 0.8)
                cur = time.time()
            else:
                break
        else:
            raise TimeoutError("wifi connect timeout.")

    def connect_wifi(self):
        for i in range(2):
            self.s20i_d.press('home')
        self.s20i_d(resourceId="com.android.systemui:id/recent_apps").click()
        if self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").exists(3):
            self.s20i_d(resourceId="com.android.systemui:id/clear_all_button").click()
            log.info("cleared background app success.")
        else:
            log.info('no background app running.')
        self.s20i_d.press('home')
        self.s20i_d(text="设置").click()
        self.s20i_d(resourceId="android:id/title", text="网络和互联网").click()
        self.s20i_d(resourceId="android:id/title", text="WLAN").click()
        if self.s20i_d(resourceId="com.android.settings:id/switch_widget").get_text() == "关闭":  # 判断wifi是否开启
            self.s20i_d(resourceId="com.android.settings:id/switch_widget").click()
        time.sleep(3)
        self.search_wifi()
        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
            return
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接")
            return
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
            return
        self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).click()  # 点击名称为“GlocalMe”的wifi
        if self.s20i_d(resourceId="com.android.settings:id/password").exists(3):  # 判断是否弹出输入password的弹窗
            self.s20i_d.send_keys(U3_settings.test_device_info['password'], clear=True)  # 输入password
            self.s20i_d(resourceId="android:id/button1").click()
        elif self.s20i_d(text="取消保存").exists(3):
            log.info("auxiliary device connect test device wifi successfully")
            return
        if self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="登录到网络").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:登录到网络")
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接")
        elif self.s20i_d(resourceId="android:id/title", text=U3_settings.test_device_info['ssid']).sibling(resourceId="android:id/summary", text="已连接，但无法访问互联网").exists(3):
            log.info("auxiliary device connect test device wifi successfully, text:已连接，但无法访问互联网")
        else:
            log.error("auxiliary device connect test device wifi failed.")
            screenShot(self.s20i_d, title="wifi_connect_fail")
            raise ConnectionError("wifi connect error.")

    def open_webui(self):
        self.connect_wifi()
        self.s20i_d(resourceId="com.android.systemui:id/center_group").click()  # 返回主界面
        self.s20i_d(text="Chrome").click()  # 点击打开浏览器
        # 判断浏览器是否第一次打开，如果第一次打开就点击设置欢迎页弹窗
        if self.s20i_d(resourceId="com.android.chrome:id/terms_accept").exists(5):
            try:
                self.s20i_d(resourceId="com.android.chrome:id/terms_accept").click()
                self.s20i_d(resourceId="com.android.chrome:id/negative_button").click()
                self.s20i_d(resourceId="com.android.chrome:id/button_secondary").click()
            except Exception as e:
                log.error(e)
                # log.info('浏览器不是第一次启动，不需要设置欢迎页')
        # self.s20i_d(resourceId="com.android.chrome:id/home_button").click()  # 点击浏览器主页按钮进入主页
        try:
            self.s20i_d(resourceId="com.android.chrome:id/url_bar").click()  # 点击浏览器搜索框
        except:
            self.s20i_d(resourceId="com.android.chrome:id/search_box").click()  # 点击浏览器搜索框
        self.s20i_d.send_keys("192.168.43.1", clear=True)  # 在搜索框中输入192.168.43.1
        self.s20i_d.xpath('//android.widget.ListView/android.view.ViewGroup[1]').click()  # 点击进入webui界面

    def login_webui(self):
        if self.s20i_d(resourceId="tr_manage_my_device").exists(5):
            self.s20i_d(resourceId="tr_manage_my_device").click()  # 点击管理我的设备
        elif self.s20i_d(text="Manage My Device").exists(5):
            self.s20i_d(text="Manage My Device").click()
        # 输入账号密码登陆管理界面
        self.s20i_d(resourceId="username").click()
        time.sleep(2)
        self.s20i_d.click(0.342, 0.303)
        self.s20i_d.send_keys("admin", clear=True)
        for i in range(2):
            self.s20i_d(resourceId="passWord").click()
        self.s20i_d.send_keys("admin", clear=True)
        self.s20i_d(text="Login").click()

if __name__ == '__main__':
    # os.environ.update({"__COMPAT_LAYER": "RUnAsInvoker"})
    current_time = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time()))
    # 创建测试记录保存路径
    test_record_path = '/'.join([BASEDIR, 'TestRecord', project_name, current_time])
    if not os.path.exists(test_record_path):
        os.makedirs(test_record_path)
    # 报告保存路径
    report_path = '/'.join([BASEDIR, 'reports', project_name])
    if not os.path.exists(report_path):
        make_report_path = os.makedirs(report_path)
    report_name = '/'.join([report_path, '.'.join([current_time, 'html'])])
    report = ''.join(['--html=', report_name])

    '''生成html报告的运行方式'''
    pytest.main(['-s', report, 'test_project_U3.py::Test_CloudSimAndPhysicalSimSwitch::test08'])

    '''生成定制测试报告的运行方式'''
    # result_file = '/'.join([report_path, time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time()))])
    # report_file = '/'.join([result_file, 'report'])
    # os.makedirs(report_file)
    # pytest.main(['-s', '-q', '--alluredir', result_file, 'test_project_S20i_RedLineCase.py::Test_Camera'])
    # cmd = ' '.join(['allure generate', result_file, '-o', report_file])
    # subprocess.getoutput(cmd)
