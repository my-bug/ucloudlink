#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'Curry'
'''
description:
    1、脚本目的：G4S项目自动化测试
    2、涵盖测试模块：
        - FOTA功能
'''
import pytest, subprocess, time, os, allure, sys
# 获取项目路径
BASEDIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
sys.path.append(BASEDIR)
from common import models, logger, appium_driver
from common.ui_operation import G4S_uiOpration
from config import settings, driver_config
from config.config import global_config
from appium import webdriver
from openpyxl import load_workbook

# 启动切口工具
subprocess.getstatusoutput(" ".join(["start", settings.Qualcomm_SwitchCom_file]))

# 获取项目名称
project_name = os.path.basename(__file__).split('.')[0]
# 创建log对象
log_name = '.'.join(['_'.join([project_name, time.strftime('%Y%m%d%H%M%S', time.localtime())]), "log"])
log_file = os.path.join(BASEDIR, "logs", log_name)
log = logger.Logger(screen_output=True, log_level='logging.INFO', log_file=log_file).create_logger()

class Test_Flash:
    def setup_class(self):
        log.info("-" * 20 + "Start Test_Flash" + "-" * 20)
        # 测试截图、视频的保存路径；如果不存在此路径，就创建此路径
        self.test_record_path = '/'.join([BASEDIR, 'TestRecord', project_name, "Test_Fota"])
        if not os.path.exists(self.test_record_path):
            os.makedirs(self.test_record_path)

        # 初始化信息
        self.serialno = global_config.getRaw("config-G4S", "serialno")
        self.desired_caps = {
            'platformName': 'Android',
            'platformVersion': '8.1.0',
            # 'appPackage': 'com.android.settings',
            # 'appActivity': 'com.android.settings.Settings',
            'newCommandTimeout': '360',
            'deviceName': self.serialno,
            'noReset': False,
        }
        self.source_version = global_config.getRaw("config-G4S", "source_version")
        self.target_version = global_config.getRaw("config-G4S", "target_version")
        self.g4s = models.G4S(device_id=self.serialno, log_project=log)
        self.g4s.wait_device_connect()
        self.g4s.wakey()  # 设置屏幕常亮

        # 安装appium服务相关的apk
        self.g4s.apk_install(settings.APPIUM_APK_FILE)
        self.g4s.apk_install(settings.UIAUTOMATOR2_APK_FILE_01)
        self.g4s.apk_install(settings.UIAUTOMATOR2_APK_FILE_02)

        # 启动appium server服务
        # self.driver = appium_driver.Appium_driver(driver_config.G4S_CONFIG).get_driver()
        self.driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", self.desired_caps)

        self.g4s_ui = G4S_uiOpration(g4s=self.g4s, driver=self.driver, log_project=log)

        self.driver.press_keycode(3)
        self.version = self.g4s.get_current_version()
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)  # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path,
                                                  '.'.join([self.version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', 'G4自动化测试用例.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["fota"]  # 选择表

        """定义界面元素变量"""
        # SIM卡网络开关的element
        self.sim_switch_element = self.driver.find_element_by_id("com.glocalme.g4home:id/st_sim_net")
        # 升级弹窗提示的element
        self.new_version_notice_element = self.driver.find_element_by_id("com.abupdate.fota_demo_iot:id/tv_new_version")

    '''测试用例：云卡网络-开机检测'''
    def test01(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.g4s.root()
        self.g4s_ui.swith_network(swith_to_sim=False)
        self.g4s.reboot()
        self.g4s.wait_device_connect()
        self.g4s.wakey()
        self.g4s.wait_network_connect()
        self.new_version_notice_element.is_displayed()









"""
import pytest, subprocess, time, os
from common import appium_driver, models, logger
from config import settings, driver_config
import uiautomator2 as u2
from pywinauto import application

# 创建log对象
log = logger.Logger(screen_output=True, log_level='logging.INFO').create_logger()
# 获取项目路径
BASEDIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
# 创建截图保存路径
test_image_path = '/'.join([BASEDIR, 'TestRecord', 'test_project_G4S'])
if not os.path.exists(test_image_path):
    make_image_path = os.makedirs(test_image_path)
# 报告保存路径
report_path = '/'.join([BASEDIR, 'reports', 'test_project_G4S'])
if not os.path.exists(report_path):
    make_report_path = os.makedirs(report_path)
report_name = '/'.join([report_path, '.'.join([time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time())), 'html'])])
report = ''.join(['--html=', report_name])


class Test_FOTA:
    # 测试类起始函数
    def setup_class(self):
        self.g4s = models.G4S(log_project=log, device_id='11c9b534')
        self.g4s.wait_device_connect()
        # 设置当设备usb连接时屏幕常亮
        self.g4s.wakey()
        # 安装获取root权限的apk
        self.g4s.apk_install(settings.G4S_ROOT_APK_FILE)
        # 安装appium服务相关的apk
        self.g4s.apk_install(settings.APPIUM_APK_FILE)
        self.g4s.apk_install(settings.UIAUTOMATOR2_APK_FILE_01)
        self.g4s.apk_install(settings.UIAUTOMATOR2_APK_FILE_02)
        # 启动appium server服务
        self.driver = appium_driver.Appium_driver(driver_config.G4S_CONFIG).get_driver()
        # 获取root状态并开启root权限
        self.driver.start_activity(app_package='com.ukl.factory', app_activity='com.ukl.factory.UklFacMainActivity')
        root_state = self.driver.find_element_by_xpath("/hierarchy/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout[2]/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.ListView/android.widget.LinearLayout[2]/android.widget.LinearLayout/android.widget.CheckBox").get_attribute('checked')
        log.info("root state:%s" % root_state)
        if root_state == 'false':
            self.driver.find_element_by_xpath("/hierarchy/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout[2]/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.ListView/android.widget.LinearLayout[2]/android.widget.LinearLayout/android.widget.CheckBox").click()
        time.sleep(2)

    def setup(self):
        self.driver.start_activity(app_package='com.glocalme.g4home',
                                   app_activity='com.glocalme.basic.more.MoreActivity')
        time.sleep(1)
        self.driver.find_element_by_android_uiautomator("text(\"SIM卡管理\")").click()

    # def teardown(self):
    #     self.driver.start_activity(app_package='com.glocalme.g4home', app_activity='com.glocalme.basic.more.MoreActivity')
    #     time.sleep(1)
    #     self.driver.find_element_by_android_uiautomator("text(\"SIM卡管理\")").click()

    def test_01(self):
        self.driver.find_element_by_android_uiautomator("text(\"升级\")").click()
        time.sleep(1)


def version_download():
    # 启动G4S刷机软件
    app = application.Application().start('F:\项目\G4S\G4S_UCloudlink_Qua_Download Tool_V1.0.0\\Ucloud_Qualcomm_DownLoad.exe')
    # app['Ucloudlink Qua_Download Version:1.0.0'].print_control_identifiers()
    g4s = models.G4S(device_id='11c9b534')
    if g4s.connect_device():
        # 发送命令让设备进入9008，然后使用刷机工具自动刷机
        subprocess.getoutput('adb reboot edl')
    g4s.wait_device_disconnect()
    g4s.wait_device_connect()
    app.kill()
    time.sleep(150)


@pytest.fixture()
def upgrade():
    version_download()

class Test_FOTA_Alternative:

    #    自定义截图函数
    def screenShot(self, title):
        screenShot_path = '/'.join([test_image_path, '.'.join(['_'.join([title, time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time()))]), 'png'])])
        log.info("screenShot_path:%s" % screenShot_path)
        try:
            self.d.screenshot(screenShot_path)
        except Exception as e:
            log.error(e)

    def fota_check(self):
        self.d(text='升级').click()
        time.sleep(0.5)
        self.d(resourceId="com.glocalme.g4home:id/tv_to_update").click()
        check_result = self.d(text='下载', resourceId="com.abupdate.fota_demo_iot:id/tv_update_state_detail").exists(30)
        return check_result

    def fota_download(self):
        if self.d(text='下载', resourceId="com.abupdate.fota_demo_iot:id/tv_update_state_detail").exists():
            self.d(resourceId="com.abupdate.fota_demo_iot:id/tv_update_state_detail").click()
        download_result = self.d(text='点击升级', resourceId="com.abupdate.fota_demo_iot:id/tv_update_state_detail").exists(120)
        return download_result

    def fota_upgrade(self):
        new_version = self.d(resourceId="com.abupdate.fota_demo_iot:id/iot_version_detail").get_text().strip("最新版本:").strip()
        log.info('detected upgrade target version:%s' % new_version)
        if self.d(text='点击升级', resourceId="com.abupdate.fota_demo_iot:id/tv_update_state_detail").exists():
            self.d(resourceId="com.abupdate.fota_demo_iot:id/image_arrow").click()
        self.g4s.wait_device_disconnect()
        self.g4s.wait_device_connect()
        self.g4s.wakey()
        self.g4s.wait_network_connect()
        time.sleep(10)
        self.d.healthcheck()
        upgraded_version = self.d.shell('getprop ro.fota.version')[0].strip()
        log.info('upgraded version:%s' % upgraded_version)
        upgrade_result = self.d(resourceId="com.abupdate.fota_demo_iot:id/tv_result", text='恭喜您，升级成功！').exists(180)
        if not upgrade_result:
            if upgraded_version == new_version:
                log.info('upgraded version same as detected target version, upgrade success.')
                upgrade_result = True
        return upgrade_result

    # 测试类起始函数
    def setup_class(self):
        self.g4s = models.G4S(log_project=log, device_id='11c9b534')
        self.g4s.wait_device_connect()
        # 设置当设备usb连接时屏幕常亮
        self.g4s.wakey()
        try:
            # 执行init，在测试设备端安装uiautomator app，minicap和minitouch，atx-agent
            uiautomator2_init = subprocess.getoutput("python -m uiautomator2 init")
            log.info("uiautomator2_init:%s" % uiautomator2_init)
        except Exception as e:
            log.error(e)
        # 手动安装uiautomator2服务相关的apk
        self.g4s.apk_install(settings.APPUIAUTOMATOR_APK_FILE)
        self.g4s.apk_install(settings.APPUIAUTOMATORTEST_APK_FILE)
        time.sleep(5)
        # 与测试设备端建立连接
        # self.d = u2.connect('11c9b534')
        retry_num = 0
        for i in range(3):
            if retry_num < 3:
                try:
                    self.d = u2.connect('11c9b534')
                    break
                except Exception as e:
                    log.error(e)
                    retry_num += 1
                    log.error("uiautomator2 connect retry num:%s" % retry_num)
            else:
                raise ConnectionError
        self.d.press('back')
        # 如果弹窗提示有新版本升级，点击取消按钮
        if self.d(resourceId="com.abupdate.fota_demo_iot:id/tv_new_version").exists():
            self.d(resourceId="com.abupdate.fota_demo_iot:id/tv_click_cancel").click()
        self.d.swipe(0.5, 0.8, 0.5, 0.5)
        time.sleep(1)
        self.d(resourceId="com.glocalme.g4home:id/tv_language", text="中文简体").click()
        self.d(resourceId="com.glocalme.g4home:id/iv_complete").click()
        self.d.swipe(0.5, 0.8, 0.5, 0.5)
        time.sleep(1)
        # 如果刷机后首次开机弹出VSIM和SIM1的选择界面，选择VSIM登录
        if self.d(resourceId="com.glocalme.g4home:id/tv_sim_detect_status").exists():
            self.d(text="VSIM").click()
            self.d(text="确定").click()
            self.g4s.wait_network_connect()
            time.sleep(60)
        while not self.d(resourceId="com.android.systemui:id/network_type").exists(30):
            log.error('network not OK!')
            self.d(resourceId="com.glocalme.g4home:id/fl_container").click()
            self.d.press('back')
        # 如果弹窗提示有新版本升级，点击取消按钮
        if self.d(resourceId="com.abupdate.fota_demo_iot:id/tv_new_version").exists(30):
            self.d(resourceId="com.abupdate.fota_demo_iot:id/tv_click_cancel").click()
        self.d(resourceId="com.glocalme.g4home:id/fl_container").click()
        self.d(resourceId="com.glocalme.g4home:id/iv_more").click()
        time.sleep(0.5)



    # 云卡网络下fota检测功能测试
    def test_01_cloudsim_check(self):
        self.d(text='SIM卡管理').click()
        sim_state = self.d(resourceId="com.glocalme.g4home:id/st_sim_net").get_text()
        log.info('SIM卡网络状态：%s' % sim_state)
        if sim_state == '开启':
            self.d(resourceId="com.glocalme.g4home:id/st_sim_net").click()
            time.sleep(1)
        self.d(resourceId="com.glocalme.g4home:id/tv_back").click()
        # self.g4s.wait_network_connect()
        while not self.d(resourceId="com.android.systemui:id/network_type").exists(30):
            log.error('network not OK!')
            self.d(text='SIM卡管理').click()
            self.d.press('back')
        check_result = self.fota_check()
        log.info("test_01_CloudsimCheck check_result:%s" % check_result)
        pytest.assume(check_result is True)
        if not check_result:
            log.error("cloudsim check failed.")
            self.screenShot(title='cloudsim_check_Failed')

    # 云卡网络下fota下载功能测试
    def test_02_cloudsim_download(self):
        download_result = self.fota_download()
        log.info("test_02_cloudsim_download result:%s" % download_result)
        pytest.assume(download_result is True)
        if not download_result:
            log.error("cloudsim download failed.")
            self.screenShot(title='cloudsim_download_failed')

    def test_03_cloudsim_upgrade(self):
        upgrade_result = self.fota_upgrade()
        log.info("test_03_cloudsim_upgrade result:%s" % upgrade_result)
        pytest.assume(upgrade_result is True)
        if not upgrade_result:
            log.error("cloudsim upgrade failed.")
            self.screenShot(title='cloudsim_upgrade_failed')

    # 实体卡网络下fota检测功能测试
    @pytest.mark.usefixtures("upgrade")
    def test_04_SIM_check(self):
        self.setup_class()
        self.d(text='SIM卡管理').click()
        sim_state = self.d(resourceId="com.glocalme.g4home:id/st_sim_net").get_text()
        log.info('SIM卡网络状态：%s' % sim_state)
        if sim_state == '关闭':
            self.d(resourceId="com.glocalme.g4home:id/st_sim_net").click()
            self.d(text="连接SIM卡网络需断开当前设备套餐，是否确认切换到SIM卡网络？").exists(3)
            self.d(text="确定").click()
        self.d(resourceId="com.glocalme.g4home:id/tv_back").click()
        while not self.d(resourceId="com.android.systemui:id/network_type").exists(30):
            log.error('network not OK!')
            self.d(text='SIM卡管理').click()
            self.d.press('back')
        check_result = self.fota_check()
        log.info("test_04_SIM_check check_result:%s" % check_result)
        pytest.assume(check_result is True)
        if not check_result:
            log.error("SIM check failed.")
            self.screenShot(title='SIM_check_Failed')
            
    def test_05_SIM_download(self):
        download_result = self.fota_download()
        log.info("test_05_SIM_download result:%s" % download_result)
        pytest.assume(download_result is True)
        if not download_result:
            log.error("SIM download failed.")
            self.screenShot(title='SIM_download_failed')
            
    def test_06_SIM_upgrade(self):
        upgrade_result = self.fota_upgrade()
        log.info("test_06_SIM_upgrade result:%s" % upgrade_result)
        pytest.assume(upgrade_result is True)
        if not upgrade_result:
            log.error("SIM upgrade failed.")
            self.screenShot(title='SIM_upgrade_failed')

    # wifi网络下fota检测功能测试
    @pytest.mark.usefixtures("upgrade")
    def test_07_wifi_check(self):
        self.setup_class()
        self.d(resourceId="com.glocalme.g4home:id/tv_item_more", text="升级").click()
        # 打开Wi-Fi开关
        self.d(resourceId="com.glocalme.g4home:id/st_update").click()
        # 连接指定的wifi
        if self.d(resourceId="com.glocalme.g4home:id/tv_item_wifi_list_ssid", text=settings.WIFI_SSID).exists(60):
            self.d(resourceId="com.glocalme.g4home:id/tv_item_wifi_list_ssid", text=settings.WIFI_SSID).click()
            self.d(resourceId="com.glocalme.g4home:id/tv_input_password").click()
            if self.d(resourceId="com.glocalme.g4home:id/et_wifiitem_pwd").exists(10):
                self.d.send_keys(settings.WIFI_PWD, clear=True)
                time.sleep(1)
            self.d(text="确定").click()
            time.sleep(1)
            self.d(text="确认").click()
        self.g4s.wait_network_connect()
        self.d(resourceId="com.glocalme.g4home:id/tv_back").click()
        self.d(resourceId="com.glocalme.g4home:id/tv_to_update").click()
        check_result = self.d(text='下载', resourceId="com.abupdate.fota_demo_iot:id/tv_update_state_detail").exists(30)
        log.info("test_07_wifi_check check_result:%s" % check_result)
        pytest.assume(check_result is True)
        if not check_result:
            log.error("wifi check failed.")
            self.screenShot(title='wifi_check_Failed')

    def test_08_wifi_download(self):
        download_result = self.fota_download()
        log.info("test_08_wifi_download result:%s" % download_result)
        pytest.assume(download_result is True)
        if not download_result:
            log.error("wifi download failed.")
            self.screenShot(title='wifi_download_failed')

    def test_09_wifi_upgrade(self):
        upgrade_result = self.fota_upgrade()
        log.info("test_09_wifi_upgrade result:%s" % upgrade_result)
        pytest.assume(upgrade_result is True)
        if not upgrade_result:
            log.error("wifi upgrade failed.")
            self.screenShot(title='wifi_upgrade_failed')

def main():
    k3p = models.G4S(log_project=log)
    for i in range(100):
        print("-"*20 + "第%d次刷机" % (i+1) + "-"*20)
        # k3p.wait_device_connect()
        if k3p.connect_device():
            # 发送命令让设备进入9008，然后使用刷机工具自动刷机
            subprocess.getoutput('adb reboot edl')
        k3p.wait_device_disconnect()
        k3p.wait_device_connect(timeout=360)
        time.sleep(120)
        # k3p.wait_network_connect()


if __name__ == '__main__':
    # version_download()
    # pytest.main(['-s', report, 'test_project_G4S.py::Test_FOTA_Alternative'])
    # print(os.path.dirname(__file__))
    main()
"""