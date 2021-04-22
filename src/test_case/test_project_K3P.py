#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'Curry'
'''
description:
    一、脚本目的：K3P项目自动化测试
    二、涵盖测试模块：
        1、
'''

import pytest, subprocess, time, os, allure, sys, random
# 获取项目路径
BASEDIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
sys.path.append(BASEDIR)
from common import models, logger
# from common.ui_operation import Operate
from config import settings, K3P_settings
import uiautomator2 as u2
# from pywinauto import application
from openpyxl import load_workbook

# 获取项目名称
project_name = os.path.basename(__file__).split('.')[0]
# 创建log对象
log_name = '.'.join(['_'.join([project_name, time.strftime('%Y%m%d%H%M%S', time.localtime())]), "log"])
log_file = os.path.join(BASEDIR, "logs", log_name)
log = logger.Logger(screen_output=True, log_level='logging.INFO', log_file=log_file).create_logger()

def screenShot(d, title):
    '''
    在uiautomator2的截图函数上封装的截图方法
    :param
        d:uiautomator2对象
        title: 自定义截图的名称
    :return: 返回截图的路径
    '''
    screenShot_file = '/'.join([test_record_path, '.'.join([title, 'png'])])
    d.screenshot(screenShot_file)
    return screenShot_file

def pull_ScreenRecord(d, title):
    list = d.adb_shell('ls /sdcard/ScreenRecord/record/')
    ScreenRecord_list = list.strip().split('\n')
    file = '/'.join(['/sdcard/ScreenRecord/record', ScreenRecord_list[-1]])
    save_file = '/'.join([test_record_path, '.'.join([title, 'mp4'])])
    d.pull(src=file, dst=save_file)
    return save_file


@allure.feature("性能指标")
class Test_PerformanceIndex:
    def setup_class(self):
        log.info("-" * 20 + "Start Test_PerformanceIndex" + "-" * 20)
        self.k3p = models.G4S(device_id=K3P_settings.K3P_info['id'], log_project=log)
        self.k3p.wait_device_connect()
        self._test_file = os.path.join(BASEDIR, 'config', 'test_file.zip')    # 测试文件路径
        self._test_storage_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'test_storage_path')    # 测试pull时文件的存放路径
        if not os.path.exists(self._test_storage_path):
            os.makedirs(self._test_storage_path)
        # self._test_file_size = os.path.getsize(self._test_file)
        self._test_file_size = round((os.path.getsize(self._test_file)/1000000), 2)    # 获取测试文件大小（字节），将单位转换为MB，并保留小数点2位
        log.info("测试文件大小：%s MB" % self._test_file_size)
        version = self.k3p.get_current_version()  # 获取测试设备的版本
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)    # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path, '.'.join([version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', '整机测试内容_有屏MIFI_V1.0.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["性能指标"]  # 选择表

    # @pytest.mark.skip
    def test01_push(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        k3p_push_avg_result_list = []
        k3p_push_duration_result_list = []
        for i in range(50):
            log.info("-"*20 + "K3P设备第%d次push" % (i+1) + "-"*20)
            k3p_push_start_time = time.time()
            k3p_push = self.k3p.push(source_file=self._test_file, target_path='/sdcard/')
            k3p_push_end_time = time.time()
            if k3p_push:
                k3p_push_duration = round((k3p_push_end_time - k3p_push_start_time), 2)    # 计算push时长
                k3p_push_duration_result_list.append(k3p_push_duration)
                log.info("push duration:%s" % k3p_push_duration)
                k3p_push_speed = round((self._test_file_size / k3p_push_duration), 2)
                log.info("push speed:%s MB/S" % k3p_push_speed)    # 计算push速率
                k3p_push_avg_result_list.append(k3p_push_speed)
            else:
                log.error("push faild.")
                k3p_push_avg_result_list.append(0)
            time.sleep(1)
        log.info('k3p_push_avg_result_list:%s' % str(k3p_push_avg_result_list))
        k3p_push_success_list = self._remove_value_from_list(k3p_push_avg_result_list, 0)
        if len(k3p_push_success_list) != len(k3p_push_avg_result_list):
            k3p_push_fail_times = len(k3p_push_avg_result_list) - len(k3p_push_success_list)
            log.error("k3p push fail times: %d" % k3p_push_fail_times)
            self.ws1['H17'] = "k3p push fail times: %d" % k3p_push_fail_times
        k3p_push_avg = sum(k3p_push_success_list) / len(k3p_push_success_list)
        # self.ws1['F17'] = round(k3p_push_avg, 2)
        self.ws1['J26'] = len(k3p_push_duration_result_list)
        self.ws1['K26'] = round((sum(k3p_push_duration_result_list) / len(k3p_push_duration_result_list)), 2)

    # @pytest.mark.skip
    def test02_pull(self):
        log.info("start %s" % sys._getframe().f_code.co_name)
        k3p_pull_result_list = []
        k3p_pull_duration_result_list = []
        # 判断k3p设备/sdcard路径下是否存在test_file.zip文件，如果不存在就push一个进去
        if not self.k3p.file_exists(file='/sdcard/test_file.zip'):
            k3p_push = self.k3p.push(source_file=self._test_file, target_path='/sdcard/')
            while not k3p_push:
                k3p_push = self.k3p.push(source_file=self._test_file, target_path='/sdcard/')
        for i in range(50):
            log.info("-"*20 + "k3p设备第%d次pull" % (i+1) + "-"*20)
            k3p_pull_start_time = time.time()
            k3p_pull = self.k3p.pull(source_file="/sdcard/test_file.zip", target_path=self._test_storage_path)
            k3p_pull_end_time = time.time()
            if k3p_pull:
                k3p_pull_duration = round((k3p_pull_end_time - k3p_pull_start_time), 2)    # 计算pull时长，保留小数点后2位
                k3p_pull_duration_result_list.append(k3p_pull_duration)
                log.info("pull duration:%s" % k3p_pull_duration)
                k3p_pull_speed = round((self._test_file_size / k3p_pull_duration), 2)    # 计算pull速度，保留小数点后2位
                log.info("pull speed:%s" % k3p_pull_speed)
                k3p_pull_result_list.append(k3p_pull_speed)
            else:
                log.error("pull failed.")
                k3p_pull_result_list.append(0)
            time.sleep(1)
        log.info("k3p_pull_result_list:%s" % str(k3p_pull_result_list))
        k3p_pull_success_list = self._remove_value_from_list(k3p_pull_result_list, 0)    # 去除k3p pull结果集中为0的结果
        if len(k3p_pull_success_list) != len(k3p_pull_result_list):
            k3p_pull_fail_times = len(k3p_pull_result_list) - len(k3p_pull_success_list)    # pull 失败次数
            log.error("k3p pull failed times: %d" % k3p_pull_fail_times)
            self.ws1['M25'] = "k3p pull failed times: %d" % k3p_pull_fail_times
        k3p_pull_avg = round((sum(k3p_pull_success_list) / len(k3p_pull_success_list)), 2)
        # self.ws1['F12'] = k3p_pull_avg
        self.ws1['J25'] = len(k3p_pull_duration_result_list)
        self.ws1['K25'] = round((sum(k3p_pull_duration_result_list) / len(k3p_pull_duration_result_list)), 2)

    def test03(self):
        '''测试用例：开关机(云卡)100次'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        test_success_times = 0
        for i in range(100):
            log.info("-" * 20 + "第%d次重启测试" % (i + 1) + "-" * 20)
            self.k3p.reboot()
            time.sleep(5)
            self.k3p.wait_device_connect()
            try:
                self.k3p.wait_network_connect()
                test_success_times += 1
            except Exception as e:
                log.error(e)
            time.sleep(3)
            # login_type_01 = self.k3p.login_type()
            # test_result = pytest.assume(login_type_01 == 1)
            # if test_result:
            #     test_success_times += 1
            # else:
            #     self.ws1['K2'] = "FAIL"
                # self.ws1['L2'] = "重启测试成功次数：%d" % test_success_times
                # log.error("test_result:%s, k3p login type:%s, test success times:%d" % (test_result, test_success_times))
                # assert False
        self.ws1['J16'] = "100"
        self.ws1['K16'] = "%d%%" % test_success_times

    def test04(self):
        '''测试用例：云卡ping包'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.k3p.wait_network_connect()
        ping_result = self.k3p.ping(duration=43200, target="www.baidu.com")  # ping "www.baidu.com"
        log.info("ping_result:%s" % str(ping_result))
        packet_loss = ping_result[2]  # 丢包率
        if packet_loss > 5:
            self.ws1['K19'] = "FAIL"
            self.ws1['L19'] = "%d packets transmitted, %d received, %.2f%% packet loss" % ping_result
            assert False
        else:
            self.ws1['K19'] = "PASS"
            
    


    def teardown(self):
        # 每条case执行结束保存一下测试结果表
        self.wb.save(self.test_result_file_path)

    def teardown_class(self):
        log.info("-" * 20 + "End Test_PerformanceIndex" + "-" * 20)

    def _remove_value_from_list(self, list, value):
        for i in list:
            if i == value:
                list.remove(i)
        return list

@allure.feature("加入云卡后的稳定性")
class Test_SystemStability:
    def setup_class(self):
        log.info("-" * 20 + "Start Test_SystemStability" + "-" * 20)
        self.k3p = models.G4S(device_id=K3P_settings.K3P_info['id'], log_project=log)
        self.k3p.wait_device_connect()
        # self._test_file = os.path.join(BASEDIR, 'config', 'test_file.zip')    # 测试文件路径
        # self._test_storage_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'test_storage_path')    # 测试pull时文件的存放路径
        # if not os.path.exists(self._test_storage_path):
        #     os.makedirs(self._test_storage_path)
        # self._test_file_size = os.path.getsize(self._test_file)
        # self._test_file_size = round((os.path.getsize(self._test_file)/1000000), 2)    # 获取测试文件大小（字节），将单位转换为MB，并保留小数点2位
        # log.info("测试文件大小：%s MB" % self._test_file_size)
        version = self.k3p.get_current_version()  # 获取测试设备的版本
        self.test_result_path = os.path.join(BASEDIR, 'TestCaseAndResult', 'TestResult', project_name)
        if not os.path.exists(self.test_result_path):
            os.makedirs(self.test_result_path)    # 创建测试结果文件夹，并以测试版本命名
        self.test_result_file_path = os.path.join(self.test_result_path, '.'.join([version, 'xlsx']))  # 创建一个以测试版本号命名的Excel表
        if os.path.exists(self.test_result_file_path):
            self.wb = load_workbook(self.test_result_file_path)
        else:
            # 打开测试用例表
            self.wb = load_workbook(os.path.join(BASEDIR, 'TestCaseAndResult', 'TestCase', '整机测试内容_有屏MIFI_V1.0.xlsx'))
        self.ws = self.wb.active  # 激活表
        self.ws1 = self.wb["加入云卡后的稳定性"]  # 选择表
        
    def test11(self):
        '''测试用例：启动云卡过程关机再开机'''
        log.info("start %s" % sys._getframe().f_code.co_name)
        self.k3p.wait_network_connect()
        test_success_times = 0
        for i in range(100):
            log.info("-" * 20 + "第%d次启动云卡过程中重启测试" % (i + 1) + "-" * 20)
            self.k3p.reboot()
            time.sleep(5)
            self.k3p.wait_device_connect()
            sleep_duration = random.randint(10, 40)  # 生成一个10-40之间的随机数，将其设置为设备重启后等待时长（比拨号成功时间短）
            log.info("sleep duration:%d" % sleep_duration)
            time.sleep(sleep_duration)
            if self.k3p.connect_network():
                self.ws1['K12'] = "FAIL"
                self.ws1['L12'] = "启动云卡过程中重启测试成功次数：%d" % test_success_times
                assert False
            test_success_times += 1
        self.ws1['K12'] = "PASS"

    def teardown(self):
        # 每条case执行结束保存一下测试结果表
        self.wb.save(self.test_result_file_path)

    def teardown_class(self):
        log.info("-" * 20 + "End Test_SystemStability" + "-" * 20)

if __name__ == '__main__':
    # os.environ.update({"__COMPAT_LAYER": "RUnAsInvoker"})
    current_time = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time()))
    # 创建测试记录保存路径
    test_record_path = '/'.join([BASEDIR, 'TestRecord', project_name, current_time])
    if not os.path.exists(test_record_path):
        os.makedirs(test_record_path)
    # 报告保存路径
    report_path = '/'.join([BASEDIR, 'reports', project_name])
    if not os.path.exists(report_path):
        make_report_path = os.makedirs(report_path)
    report_name = '/'.join([report_path, '.'.join([current_time, 'html'])])
    report = ''.join(['--html=', report_name])

    '''生成html报告的运行方式'''
    pytest.main(['-s', report, 'test_project_k3p.py::Test_CloudSimAndPhysicalSimSwitch::test08'])

    '''生成定制测试报告的运行方式'''
    # result_file = '/'.join([report_path, time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time()))])
    # report_file = '/'.join([result_file, 'report'])
    # os.makedirs(report_file)
    # pytest.main(['-s', '-q', '--alluredir', result_file, 'test_project_S20i_RedLineCase.py::Test_Camera'])
    # cmd = ' '.join(['allure generate', result_file, '-o', report_file])
    # subprocess.getoutput(cmd)
